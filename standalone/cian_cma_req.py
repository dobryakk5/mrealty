import re
import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# === настройки ===
LISTING_URLS = [
    "https://www.cian.ru/sale/flat/315720090",
    "https://www.cian.ru/sale/flat/312736530",
    "https://www.cian.ru/sale/flat/319705138",
    "https://www.cian.ru/sale/flat/319629864",
    "https://www.cian.ru/sale/flat/318236887",
    "https://www.cian.ru/sale/flat/320223651",
    "https://www.cian.ru/sale/flat/319571615",
    "https://www.cian.ru/sale/flat/317667893",
    "https://www.cian.ru/sale/flat/319968025",
]

HEADERS = {
    'User-Agent': (
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
        'AppleWebKit/537.36 (KHTML, like Gecko) '
        'Chrome/115.0.0.0 Safari/537.36'
    ),
    'Accept-Language': 'ru-RU,ru;q=0.9',
}

def extract_number(text: str):
    if not text or text == '—':
        return None
    c = re.sub(r"[^\d.,]", "", text).replace('\u00A0', '').replace(' ', '').replace(',', '.')
    try:
        return float(c) if '.' in c else int(c)
    except ValueError:
        return None

def format_price(raw):
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return ''
    return f"{int(round(raw)):,}".replace(',', ' ')

def parse_listing(url: str, session: requests.Session) -> dict:
    resp = session.get(url, headers=HEADERS)
    resp.encoding = 'utf-8'
    soup = BeautifulSoup(resp.text, 'html.parser')

    data = {'URL': url}

    # Статус
    status_text = soup.find(string=re.compile(r"Объявление снято с публикации", re.IGNORECASE))
    data['Статус'] = 'Снято' if status_text else 'Активно'

    # Метки
    labels = []
    for child in soup.select('div[data-name="LabelsLayoutNew"] > span'):
        spans = child.find_all('span')
        if spans:
            lbl = spans[-1].get_text(strip=True)
            if lbl:
                labels.append(lbl)
    data['Метки'] = '; '.join(labels) if labels else None

    # Комнат
    h1 = soup.find('h1')
    if h1:
        m = re.search(r"(\d+)[^\d]*[-–]?комн", h1.get_text())
        data['Комнат'] = extract_number(m.group(1)) if m else None

    # Цена_raw
    price_el = (
        soup.select_one('div[data-name="NewbuildingPriceInfo"] [data-testid="price-amount"] span')
        or soup.select_one('div[data-name="AsideGroup"] div[data-name="PriceInfo"] [data-testid="price-amount"] span')
    )
    data['Цена_raw'] = extract_number(price_el.get_text() if price_el else None)

    # OfferSummaryInfoLayout
    summary = soup.select_one('[data-name="OfferSummaryInfoLayout"]')
    if summary:
        for item in summary.select('[data-name="OfferSummaryInfoItem"]'):
            ps = item.find_all('p')
            if len(ps) < 2:
                continue
            key = ps[0].get_text(strip=True)
            val = ps[1].get_text(strip=True)
            if key:
                if key == 'Строительная серия':
                    data[key] = val
                else:
                    data[key] = extract_number(val) if re.search(r"\d", val) else val

    # ObjectFactoids
    cont = soup.find('div', {'data-name': 'ObjectFactoids'})
    if cont:
        lines = cont.get_text(separator='\n', strip=True).split('\n')
        for i in range(0, len(lines)-1, 2):
            key = lines[i].strip()
            val = lines[i+1].strip()
            if key:
                if key == 'Строительная серия':
                    data[key] = val
                else:
                    data[key] = val

    # Просмотры
    stats_pattern = re.compile(
        r"([\d\s]+)\sпросмотр\S*,\s*(\d+)\sза сегодня,\s*(\d+)\sуникаль\S*",
        re.IGNORECASE
    )
    stats_text = soup.find(string=stats_pattern)
    if stats_text:
        m = stats_pattern.search(stats_text)
        if m:
            data['Всего просмотров'] = extract_number(m.group(1))
            data['Просмотров сегодня'] = extract_number(m.group(2))
            data['Уникальных просмотров'] = extract_number(m.group(3))

    # Адрес (убираем город Москва)
    address = None
    geo_div = soup.select_one('div[data-name="Geo"]')
    if geo_div:
        addr_span = geo_div.find('span', attrs={'itemprop': 'name'})
        if addr_span and addr_span.get('content'):
            address = addr_span['content'].strip()
        else:
            parts = [a.get_text(strip=True) for a in geo_div.select('a[data-name="AddressItem"]')]
            if parts:
                address = ', '.join(parts)
        if address and address.startswith('Москва,'):
            address = address.split(',', 1)[1].strip()
    data['Адрес'] = address

    return data

def write_styled_excel(df: pd.DataFrame, filename: str):
    df.to_excel(filename, index=False)
    wb = load_workbook(filename)
    ws = wb.active

    gray_font = Font(color="808080")
    bold_header = Font(bold=True)

    # Заголовки
    for cell in ws[1]:
        cell.font = bold_header

    # Применяем серый цвет к строкам со статусом "Снято"
    for row in range(2, ws.max_row + 1):
        status = ws[f"{get_column_letter(ws.max_column)}{row}"].value
        if status == "Снято":
            for col in range(1, ws.max_column + 1):
                ws[f"{get_column_letter(col)}{row}"].font = gray_font

    wb.save(filename)

def main():
    sess = requests.Session()
    all_rows = []
    for url in LISTING_URLS:
        print("Парсим:", url)
        try:
            all_rows.append(parse_listing(url, sess))
        except Exception as e:
            print("Ошибка при парсинге", url, ":", e)

    df = pd.DataFrame(all_rows)

    # Цена — форматируем
    if 'Цена_raw' in df.columns:
        df['Цена'] = df['Цена_raw'].apply(format_price)

    # Сортировка по цене по возрастанию
    df = df.sort_values(by='Цена_raw', ascending=True)

    # Удаляем сырой ценовой столбец
    if 'Цена_raw' in df.columns:
        df.drop(columns=['Цена_raw'], inplace=True)

    # Устанавливаем порядок колонок, адрес в конец
    base_cols = ['URL', 'Комнат', 'Цена', 'Всего просмотров', 'Просмотров сегодня', 'Уникальных просмотров']
    other_cols = [col for col in df.columns if col not in base_cols + ['Метки', 'Статус', 'Адрес']]
    final_cols = base_cols + other_cols + ['Метки', 'Статус', 'Адрес']
    df = df[[col for col in final_cols if col in df.columns]]

    # Сохраняем с форматированием
    write_styled_excel(df, 'listings.xlsx')
    print("Готово — данные сохранены в listings.xlsx")

if __name__ == '__main__':
    main()
