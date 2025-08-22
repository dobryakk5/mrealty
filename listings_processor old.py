import re, json
import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import List, Dict, Any, Tuple
import asyncio
from datetime import datetime


# Асинхронное сохранение в БД
from db_handler import save_listings, find_similar_ads_grouped, call_update_ad

# Импортируем модуль для работы с фотографиями
from photo_processor import PhotoProcessor

# Заголовки для HTTP-запросов
HEADERS = {
    'User-Agent': (
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
        'AppleWebKit/537.36 (KHTML, like Gecko) '
        'Chrome/115.0.0.0 Safari/537.36'
    ),
    'Accept-Language': 'ru-RU,ru;q=0.9',
}

class ListingsProcessor:
    """Класс для обработки объявлений недвижимости"""
    
    def __init__(self):
        self.photo_processor = PhotoProcessor()
    
    def extract_photo_urls(self, soup: BeautifulSoup) -> list[str]:
        """Извлекает ссылки на все фотографии из галереи CIAN"""
        photo_urls = []
        
        try:
            # Ищем галерею по data-name="GalleryInnerComponent"
            gallery = soup.find('div', {'data-name': 'GalleryInnerComponent'})
            if not gallery:
                return photo_urls
            
            # Ищем все изображения в галерее
            # Основные изображения
            images = gallery.find_all('img', src=True)
            for img in images:
                src = img.get('src')
                if src and src.startswith('http') and 'cdn-cian.ru' in src:
                    photo_urls.append(src)
            
            # Изображения в background-image (для видео и некоторых фото)
            elements_with_bg = gallery.find_all(style=re.compile(r'background-image'))
            for elem in elements_with_bg:
                style = elem.get('style', '')
                bg_match = re.search(r'background-image:\s*url\(["\']?([^"\')\s]+)["\']?\)', style)
                if bg_match:
                    bg_url = bg_match.group(1)
                    if bg_url.startswith('http') and ('cdn-cian.ru' in bg_url or 'kinescopecdn.net' in bg_url):
                        photo_urls.append(bg_url)
            
            # Убираем дубликаты, сохраняя порядок
            seen = set()
            unique_photos = []
            for url in photo_urls:
                if url not in seen:
                    seen.add(url)
                    unique_photos.append(url)
            
            return unique_photos
            
        except Exception as e:
            print(f"Ошибка при извлечении фотографий: {e}")
            return []
    
    async def extract_photo_urls_from_url(self, url: str) -> list[str]:
        """Асинхронно получает URL и извлекает ссылки на фотографии"""
        try:
            # Используем более надежный способ для асинхронных запросов
            loop = asyncio.get_event_loop()
            response = await loop.run_in_executor(None, lambda: requests.get(url, headers=HEADERS))
            response.raise_for_status()
            soup = BeautifulSoup(response.text, 'html.parser')
            return self.extract_photo_urls(soup)
        except Exception as e:
            print(f"Ошибка при получении URL {url}: {e}")
            return []
    
    def extract_listing_info(self, url: str) -> dict:
        """Извлекает основную информацию об объявлении"""
        try:
            response = requests.get(url, headers=HEADERS)
            response.raise_for_status()
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # Парсим основную информацию
            listing_data = parse_listing(url, requests.Session())
            
            # Формируем структурированную информацию
            info = {
                'rooms': listing_data.get('Комнат', 'N/A'),
                'price': listing_data.get('Цена_raw', 'N/A'),
                'floor': listing_data.get('Этаж', 'N/A'),
                'total_area': listing_data.get('Общая площадь', 'N/A'),
                'kitchen_area': listing_data.get('Площадь кухни', 'N/A'),
                'metro': listing_data.get('Минут метро', 'N/A')
            }
            
            return info
        except Exception as e:
            print(f"Ошибка при извлечении информации об объявлении {url}: {str(e)}")
            return {
                'rooms': 'N/A',
                'price': 'N/A',
                'floor': 'N/A',
                'total_area': 'N/A',
                'kitchen_area': 'N/A',
                'metro': 'N/A'
            }
    
    def generate_html_gallery(self, listing_urls: list[str], user_id: int, subtitle: str = None) -> str:
        """Генерирует HTML галерею с внешними ссылками на фотографии"""
        html_parts = []
        
        html_parts.append("""
        <!DOCTYPE html>
        <html lang="ru">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Подбор недвижимости</title>
            <style>
                body { font-family: Arial, sans-serif; margin: 20px; background-color: #f5f5f5; }
                .listing { background: white; margin: 20px 0; padding: 20px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
                .listing h3 { color: #333; margin-top: 0; margin-bottom: 15px; }
                .listing p { margin: 8px 0; color: #555; }
                .listing strong { color: #333; }
                .main-title { color: #333; margin-bottom: 10px; }
                .subtitle { color: #666; font-size: 18px; margin-bottom: 30px; font-style: italic; }
                .photo-grid { 
                    display: grid; 
                    grid-template-columns: repeat(auto-fill, minmax(180px, 1fr)); 
                    gap: 8px; 
                    margin: 15px 0; 
                    max-height: 600px; 
                    overflow-y: auto; 
                    padding: 10px;
                    border: 1px solid #eee;
                    border-radius: 8px;
                }
                .photo-item { position: relative; }
                .photo-item img { 
                    width: 100%; 
                    height: 140px; 
                    object-fit: cover; 
                    border-radius: 5px; 
                    border: 2px solid transparent;
                    transition: border-color 0.2s;
                }
                .photo-item img:hover { 
                    border-color: #0066cc;
                }
                .no-photos { color: #666; font-style: italic; }
                
                /* Мобильная адаптация */
                @media (max-width: 768px) {
                    body { margin: 10px; }
                    .listing { padding: 15px; margin: 15px 0; }
                    .photo-grid { 
                        grid-template-columns: repeat(auto-fill, minmax(150px, 1fr)); 
                        gap: 6px; 
                        padding: 8px;
                    }
                    .photo-item img { 
                        height: 120px; 
                    }
                    .main-title { font-size: 24px; }
                    .subtitle { font-size: 16px; }
                }
            </style>
        </head>
        <body>
            <h1 class="main-title">🏠 Подбор недвижимости</h1>
        """)
        
        if subtitle:
            html_parts.append(f'<h2 class="subtitle">{subtitle}</h2>')
        
        html_parts.append("")
        
        for i, listing_url in enumerate(listing_urls, 1):
            try:
                # Парсим объявление
                listing_data = parse_listing(listing_url, requests.Session())
                
                # Извлекаем фотографии
                soup = BeautifulSoup(requests.get(listing_url, headers=HEADERS).text, 'html.parser')
                photo_urls = self.extract_photo_urls(soup)
                
                html_parts.append(f"""
                <div class="listing">
                    <h3>Вариант #{i}</h3>
                """)
                
                # Добавляем основную информацию
                if 'Комнат' in listing_data and listing_data['Комнат']:
                    html_parts.append(f"<p><strong>Комнат:</strong> {listing_data['Комнат']}</p>")
                if 'Цена_raw' in listing_data and listing_data['Цена_raw']:
                    html_parts.append(f"<p><strong>Цена:</strong> {listing_data['Цена_raw']:,} ₽</p>")
                
                # Добавляем этаж/этажность
                if 'Этаж' in listing_data and listing_data['Этаж']:
                    html_parts.append(f"<p><strong>Этаж:</strong> {listing_data['Этаж']}</p>")
                
                # Добавляем метраж общий
                if 'Общая площадь' in listing_data and listing_data['Общая площадь']:
                    html_parts.append(f"<p><strong>Общая площадь:</strong> {listing_data['Общая площадь']} м²</p>")
                
                # Добавляем кухню
                if 'Площадь кухни' in listing_data and listing_data['Площадь кухни']:
                    html_parts.append(f"<p><strong>Кухня:</strong> {listing_data['Площадь кухни']} м²</p>")
                
                # Переименовываем "Метро" в "Минут до метро"
                if 'Минут метро' in listing_data and listing_data['Минут метро']:
                    html_parts.append(f"<p><strong>Минут до метро:</strong> {listing_data['Минут метро']}</p>")
                
                # Добавляем фотографии
                if photo_urls:
                    # Генерируем сетку фотографий (все фото без ограничений)
                    html_parts.append(f'<div class="photo-grid">')
                    for j, photo_url in enumerate(photo_urls):
                        html_parts.append(f"""
                        <div class="photo-item">
                            <img src="{photo_url}" alt="Фото {j+1}" 
                                 onerror="this.style.display='none'; this.nextElementSibling.style.display='block';"
                                 loading="lazy">
                            <div class="photo-fallback" style="display: none; background: #f0f0f0; border: 1px dashed #ccc; border-radius: 5px; padding: 20px; text-align: center; color: #666;">
                                <div>📷 Фото {j+1}</div>
                                <div style="font-size: 12px; margin-top: 5px;">
                                    <a href="{photo_url}" target="_blank" style="color: #0066cc;">Открыть фото</a>
                                </div>
                            </div>
                        </div>
                        """)
                    html_parts.append('</div>')
                else:
                    html_parts.append('<p class="no-photos">📷 Фотографии не найдены</p>')
                
                html_parts.append('</div>')
                
            except Exception as e:
                html_parts.append(f"""
                <div class="listing">
                    <h3>Вариант #{i}</h3>
                    <p style="color: red;">Ошибка при парсинге: {str(e)}</p>
                </div>
                """)
        
        html_parts.append("""
        </body>
        </html>
        """)
        
        return ''.join(html_parts)
    
    async def generate_html_gallery_embedded(self, listing_urls: list[str], user_id: int, subtitle: str = None, remove_watermarks: bool = False) -> str:
        """Генерирует HTML галерею с встроенными Base64 изображениями"""
        html_content = f"""
        <!DOCTYPE html>
        <html lang="ru">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Подбор недвижимости</title>
            <style>
                body {{ 
                    font-family: Arial, sans-serif; 
                    margin: 20px; 
                    background-color: #f5f5f5; 
                }}
                .listing {{ 
                    background: white; 
                    margin: 20px 0; 
                    padding: 20px; 
                    border-radius: 10px; 
                    box-shadow: 0 2px 10px rgba(0,0,0,0.1); 
                }}
                .listing h3 {{ 
                    color: #333; 
                    margin-top: 0; 
                    margin-bottom: 15px; 
                }}
                .listing p {{ 
                    margin: 8px 0; 
                    color: #555; 
                }}
                .listing strong {{ 
                    color: #333; 
                }}
                .main-title {{ 
                    color: #333; 
                    margin-bottom: 10px; 
                }}
                .subtitle {{ 
                    color: #666; 
                    font-size: 18px; 
                    margin-bottom: 30px; 
                    font-style: italic; 
                }}
                .photo-grid {{ 
                    display: grid; 
                    grid-template-columns: repeat(auto-fill, minmax(180px, 1fr)); 
                    gap: 8px; 
                    margin: 15px 0; 
                    max-height: 600px; 
                    overflow-y: auto; 
                    padding: 10px;
                    border: 1px solid #eee;
                    border-radius: 8px;
                }}
                .photo-item {{ 
                    position: relative; 
                }}
                .photo-item img {{ 
                    width: 100%; 
                    height: 140px; 
                    object-fit: cover; 
                    border-radius: 5px; 
                    border: 2px solid transparent;
                    transition: border-color 0.2s;
                    background: #f8f9fa;
                }}
                .photo-item img:hover {{ 
                    border-color: #0066cc;
                }}
                .photo-fallback {{ 
                    width: 100%; 
                    height: 140px; 
                    display: flex; 
                    flex-direction: column; 
                    justify-content: center; 
                    align-items: center;
                }}
                .no-photos {{ 
                    color: #666; 
                    font-style: italic; 
                }}
                .photo-info {{ 
                    background: #f8f9fa; 
                    padding: 15px; 
                    border-radius: 8px; 
                    margin-top: 15px; 
                    font-size: 14px;
                    border-left: 4px solid #0066cc;
                }}
                .photo-info strong {{ 
                    color: #333; 
                }}
                .photo-info small {{ 
                    line-height: 1.4; 
                }}
                
                /* Мобильная адаптация */
                @media (max-width: 768px) {{
                    body {{ margin: 10px; }}
                    .listing {{ padding: 15px; margin: 15px 0; }}
                    .photo-grid {{ 
                        grid-template-columns: repeat(auto-fill, minmax(150px, 1fr)); 
                        gap: 6px; 
                        padding: 8px;
                    }}
                    .photo-item img, .photo-fallback {{ 
                        height: 120px; 
                    }}
                    .main-title {{ font-size: 24px; }}
                    .subtitle {{ font-size: 16px; }}
                }}
            </style>
        </head>
        <body>
            <h1 class="main-title">🏠 Подбор недвижимости</h1>
        """
        
        if subtitle:
            html_content += f'<h2 class="subtitle">{subtitle}</h2>'
        
        html_content += """
        """
        
        for i, listing_url in enumerate(listing_urls, 1):
            try:
                # Получаем фото для этого объявления
                photo_urls = await self.extract_photo_urls_from_url(listing_url)
                
                if photo_urls:
                    # Обрабатываем фото с поддержкой водяных знаков
                    if remove_watermarks:
                        processed_photos = self.photo_processor.process_photos_for_embedded_html(photo_urls, remove_watermarks=True)
                    else:
                        processed_photos = self.photo_processor.process_photos_for_embedded_html(photo_urls, remove_watermarks=False)
                    
                    # Извлекаем информацию об объявлении
                    listing_info = self.extract_listing_info(listing_url)
                    
                    html_content += f"""
                    <div class="listing">
                        <h3>Вариант #{i}</h3>
                        <p><strong>Комнат:</strong> {listing_info.get('rooms', 'N/A')}</p>
                        <p><strong>Цена:</strong> {listing_info.get('price', 'N/A')}</p>
                        <p><strong>Этаж:</strong> {listing_info.get('floor', 'N/A')}</p>
                        <p><strong>Общая площадь:</strong> {listing_info.get('total_area', 'N/A')}</p>
                        <p><strong>Кухня:</strong> {listing_info.get('kitchen_area', 'N/A')}</p>
                        <p><strong>Минут до метро:</strong> {listing_info.get('metro', 'N/A')}</p>
                        <div class="photo-grid">
                    """
                    
                    # Добавляем фото
                    for j, photo_data in enumerate(processed_photos, 1):
                        if photo_data and 'base64' in photo_data:
                            html_content += f"""
                            <div class="photo-item">
                                <img src="data:image/{photo_data['format']};base64,{photo_data['base64']}" 
                                     alt="Фото {j}" 
                                     loading="lazy">
                            </div>
                            """
                    
                    html_content += f"""
                        </div>
                    </div>
                    """
                else:
                    html_content += f"""
                    <div class="listing">
                        <h3>Вариант #{i}</h3>
                        <p>Фотографии не найдены</p>
                    </div>
                    """
                    
            except Exception as e:
                html_content += f"""
                <div class="listing">
                    <h3>Вариант #{i}</h3>
                    <p>Ошибка при обработке: {str(e)}</p>
                </div>
                """
        
        html_content += """
        </body>
        </html>
        """
        
        return html_content

def extract_number(text: str):
    if not text or text == '—':
        return None
    cleaned = re.sub(r"[^\d.,]", "", text)
    cleaned = cleaned.replace('\u00A0', '').replace(' ', '').replace(',', '.')
    try:
        return float(cleaned) if '.' in cleaned else int(cleaned)
    except ValueError:
        return None

async def export_listings_to_excel(listing_urls: list[str], user_id: int, output_path: str = None) -> tuple[BytesIO, int]:
    """
    Парсит список объявлений, сохраняет их в БД и возвращает Excel-файл и request_id.
    :param listing_urls: список URL объявлений
    :param user_id: ID пользователя для сохранения в БД
    :param output_path: опциональный путь для сохранения файла на диск
    :return: tuple (BytesIO с данными файла, request_id)
    """
    sess = requests.Session()
    rows = []
    for url in listing_urls:
        try:
            rows.append(parse_listing(url, sess))
        except Exception as e:
            print(f"Ошибка при парсинге {url}: {e}")

    # Сохраняем и получаем request_id
    request_id = await save_listings(rows, user_id)

    # Формируем DataFrame
    df = pd.DataFrame(rows)

    # Обработка цен
    if 'Цена_raw' in df.columns:
        df['Цена'] = df['Цена_raw']
        df = df.sort_values('Цена_raw')
        df.drop('Цена_raw', axis=1, inplace=True)

    # Порядок колонок
    ordered = [
        'Комнат', 'Цена', 'Общая площадь', 'Жилая площадь',
        'Площадь кухни', 'Санузел', 'Балкон/лоджия', 'Вид из окон',
        'Ремонт', 'Этаж', 'Год постройки', 'Строительная серия',
        'Тип дома', 'Тип перекрытий', 'Количество лифтов', 'Парковка',
        'Подъезды', 'Отопление', 'Аварийность', 'Газоснабжение',
        'Всего просмотров', 'Просмотров сегодня', 'Уникальных просмотров',
        'Адрес', 'Минут метро', 'Метки', 'Статус', 'Тип жилья', 'URL'
    ]
    df = df[[c for c in ordered if c in df.columns]]

    # Запись в BytesIO
    bio = BytesIO()
    df.to_excel(bio, index=False)
    bio.seek(0)

    # Если указан путь, сохраняем файл и форматируем столбец 'Цена'
    if output_path:
        with open(output_path, 'wb') as f:
            f.write(bio.getbuffer())

        wb = load_workbook(output_path)
        ws = wb.active
        # Выделяем заголовки
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Определяем колонку 'Цена' и задаем формат тысяч
        price_idx = df.columns.get_loc('Цена') + 1
        price_col = get_column_letter(price_idx)
        custom_format = '#,##0'
        for row in range(2, ws.max_row + 1):
            cell = ws[f"{price_col}{row}"]
            if isinstance(cell.value, (int, float)):
                cell.number_format = custom_format

        wb.save(output_path)

    return bio, request_id

# Полный парсинг страницы объявления
def parse_listing(url: str, session: requests.Session) -> dict:
    resp = session.get(url, headers=HEADERS)
    resp.encoding = 'utf-8'
    soup = BeautifulSoup(resp.text, 'html.parser')

    data = {'URL': url}
    # Определяем возможную блокировку (капча/антибот)
    page_text = soup.get_text(" ", strip=True).lower()
    is_blocked = bool(re.search(r"подтвердите, что запросы.*не робот|похожи на автоматические", page_text))
    if is_blocked:
        data['Статус'] = None
    elif soup.find(string=re.compile(r"Объявление снято", re.IGNORECASE)):
        data['Статус'] = 'Снято'
    labels = [span.get_text(strip=True) for span in soup.select('div[data-name="LabelsLayoutNew"] > span span:last-of-type')]
    data['Метки'] = '; '.join(labels) if labels else None
    h1 = soup.find('h1')
    if h1:
        m = re.search(r"(\d+)[^\d]*[-–]?комн", h1.get_text())
        if m:
            data['Комнат'] = extract_number(m.group(1))
    price_el = (
        soup.select_one('[data-name="NewbuildingPriceInfo"] [data-testid="price-amount"] span')
        or soup.select_one('[data-name="AsideGroup"] [data-testid="price-amount"] span')
    )
    if price_el:
        data['Цена_raw'] = extract_number(price_el.get_text())
        # Если статус ещё не определён и удалось найти цену — считаем объявление активным
        if 'Статус' not in data or data['Статус'] is None:
            data['Статус'] = 'Активно'

    summary = soup.select_one('[data-name="OfferSummaryInfoLayout"]')
    if summary:
        for item in summary.select('[data-name="OfferSummaryInfoItem"]'):
            ps = item.find_all('p')
            if len(ps) < 2:
                continue
            key = ps[0].get_text(strip=True)
            val = ps[1].get_text(strip=True)
            kl = key.lower().strip()
            if key == 'Строительная серия':
                data[key] = val
                continue
            if kl == 'этаж': data['Этаж'] = val; continue
            if kl in ['санузел', 'балкон/лоджия', 'количество лифтов']:
                data[key] = val; continue
            data[key] = extract_number(val) if re.search(r"\d", val) else val

    cont = soup.find('div', {'data-name': 'ObjectFactoids'})
    if cont:
        lines = cont.get_text(separator='\n', strip=True).split('\n')
        for i in range(0, len(lines)-1, 2):
            key, val = lines[i].strip(), lines[i+1].strip()
            kl = key.lower().strip()
            if key == 'Строительная серия': data[key] = val; continue
            if kl == 'этаж' and 'Этаж' not in data: data['Этаж'] = val
            elif kl in ['санузел', 'балкон/лоджия', 'количество лифтов']: data[key] = val
            else: data[key] = extract_number(val) if re.search(r"\d", val) else val

    stats_re = re.compile(r"([\d\s]+)\sпросмотр\S*,\s*(\d+)\sза сегодня,\s*(\d+)\sуникаль", re.IGNORECASE)
    st = soup.find(string=stats_re)
    if st:
        m = stats_re.search(st)
        data['Всего просмотров'], data['Просмотров сегодня'], data['Уникальных просмотров'] = (
            extract_number(m.group(1)), extract_number(m.group(2)), extract_number(m.group(3))
        )

    geo = soup.select_one('div[data-name="Geo"]')
    if geo:
        span = geo.find('span', itemprop='name')
        addr = span['content'] if span and span.get('content') else ', '.join(
            a.get_text(strip=True) for a in geo.select('a[data-name="AddressItem"]')
        )
        parts = [s.strip() for s in addr.split(',') if s.strip()]
        data['Адрес'] = ', '.join(parts[-2:]) if len(parts) > 1 else addr
        stations = []
        for li in geo.select('ul[data-name="UndergroundList"] li[data-name="UndergroundItem"]'):
            st_el = li.find('a', href=True)
            tm_el = li.find('span', class_=re.compile(r".*underground_time.*"))
            if st_el and tm_el:
                name = st_el.get_text(strip=True)
                m = re.search(r"(\d+)", tm_el.get_text(strip=True))
                stations.append((name, int(m.group(1)) if m else None))
        if stations:
            station, time_to = min(stations, key=lambda x: x[1] or float('inf'))
            data['Минут метро'] = f"{time_to} {station}"

    return data

def extract_urls(raw_input: str) -> tuple[list[str], int]:
    urls = re.findall(r'https?://[^\s,;]+', raw_input)
    return urls, len(urls)

# Создаем экземпляр класса для использования в других модулях
listings_processor = ListingsProcessor()

if __name__ == '__main__':
    user_id = 12345
    urls = ["https://www.cian.ru/sale/flat/318826533/"]
    excel = asyncio.run(export_listings_to_excel(urls, user_id, output_path="listings_numeric.xlsx"))
    print("listings_numeric.xlsx создан.")
