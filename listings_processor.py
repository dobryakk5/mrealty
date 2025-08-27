import re
import json
import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import List, Dict, Any, Tuple
import asyncio
from datetime import datetime
from selenium.webdriver.common.by import By

# Асинхронное сохранение в БД
from db_handler import save_listings, find_similar_ads_grouped, call_update_ad

# Импортируем модуль для работы с фотографиями
from photo_processor import PhotoProcessor

# Импортируем парсер Avito
try:
    from avito_parser_integration import AvitoCardParser
    AVITO_AVAILABLE = True
except ImportError:
    AVITO_AVAILABLE = False
    print("⚠️ Модуль avito_parser_integration не найден, парсинг Avito недоступен")

# Заголовки для HTTP-запросов
HEADERS = {
    'User-Agent': (
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
        'AppleWebKit/537.36 (KHTML, like Gecko) '
        'Chrome/115.0.0.0 Safari/537.36'
    ),
    'Accept-Language': 'ru-RU,ru;q=0.9',
}

class ListingsProcessor:
    """Класс для обработки объявлений недвижимости"""
    
    def __init__(self):
        self.photo_processor = PhotoProcessor()
    
    def is_avito_url(self, url: str) -> bool:
        """Определяет, является ли ссылка ссылкой на Avito"""
        return 'avito.ru' in url.lower()
    
    def is_cian_url(self, url: str) -> bool:
        """Определяет, является ли ссылка ссылкой на Cian"""
        return 'cian.ru' in url.lower()
    
    def get_url_source(self, url: str) -> int:
        """Возвращает источник ссылки: 1 - Avito, 4 - Cian"""
        if self.is_avito_url(url):
            return 1  # Avito
        elif self.is_cian_url(url):
            return 4  # Cian
        else:
            return 0  # Неизвестный источник
    
    async def parse_avito_listing(self, url: str) -> dict:
        """Парсит объявление с Avito и возвращает данные в формате для БД"""
        if not AVITO_AVAILABLE:
            print("❌ Парсер Avito недоступен")
            return None
        
        try:
            print(f"🔄 Парсим объявление Avito: {url}")
            
            # Создаем парсер Avito
            parser = AvitoCardParser()
            
            # Парсим полную страницу объявления
            parsed_data = parser.parse_avito_page(url)
            if not parsed_data:
                print("❌ Не удалось спарсить данные объявления Avito")
                return None
            
            # Преобразуем данные в формат для БД
            db_data = parser.prepare_data_for_db(parsed_data)
            if not db_data:
                print("❌ Не удалось подготовить данные для БД")
                return None
            
            # Добавляем источник
            db_data['source'] = 1  # Avito
            
            print(f"✅ Объявление Avito успешно спарсено")
            return db_data
            
        except Exception as e:
            print(f"❌ Ошибка парсинга объявления Avito: {e}")
            return None
        finally:
            # Закрываем браузер
            if 'parser' in locals() and parser.driver:
                try:
                    parser.cleanup()
                except:
                    pass
    
    async def parse_listing_universal(self, url: str) -> dict:
        """Универсальный метод для парсинга объявлений с Cian и Avito"""
        try:
            if self.is_avito_url(url):
                print(f"🏠 Парсим объявление Avito: {url}")
                return await self.parse_avito_listing(url)
            elif self.is_cian_url(url):
                print(f"🏠 Парсим объявление Cian: {url}")
                # Используем существующий метод для Cian
                session = requests.Session()
                return parse_listing(url, session)
            else:
                print(f"⚠️ Неизвестный источник ссылки: {url}")
                return None
        except Exception as e:
            print(f"❌ Ошибка универсального парсинга: {e}")
            return None
    
    def extract_photo_urls(self, soup: BeautifulSoup) -> list[str]:
        """Извлекает ссылки на все фотографии из галереи CIAN"""
        photo_urls = []
        
        try:
            # Ищем галерею по data-name="GalleryInnerComponent"
            gallery = soup.find('div', {'data-name': 'GalleryInnerComponent'})
            if not gallery:
                return photo_urls
            
            # Ищем все изображения в галерее
            # Основные изображения
            images = gallery.find_all('img', src=True)
            for img in images:
                src = img.get('src')
                if src and src.startswith('http') and 'cdn-cian.ru' in src:
                    photo_urls.append(src)
            
            # Изображения в background-image (для видео и некоторых фото)
            elements_with_bg = gallery.find_all(style=re.compile(r'background-image'))
            for elem in elements_with_bg:
                style = elem.get('style', '')
                bg_match = re.search(r'background-image:\s*url\(["\']?([^"\')\s]+)["\']?\)', style)
                if bg_match:
                    bg_url = bg_match.group(1)
                    if bg_url.startswith('http') and ('cdn-cian.ru' in bg_url or 'kinescopecdn.net' in bg_url):
                        photo_urls.append(bg_url)
            
            # Убираем дубликаты, сохраняя порядок
            seen = set()
            unique_photos = []
            for url in photo_urls:
                if url not in seen:
                    seen.add(url)
                    unique_photos.append(url)
            
            return unique_photos
            
        except Exception as e:
            print(f"Ошибка при извлечении фотографий: {e}")
            return []
    
    def extract_avito_photo_urls(self, soup: BeautifulSoup) -> list[str]:
        """Извлекает ссылки на фотографии с Avito"""
        photo_urls = []
        
        try:
            # Ищем галерею фотографий на Avito
            # Основные изображения в карточке
            images = soup.find_all('img', src=True)
            for img in images:
                src = img.get('src')
                if src and src.startswith('http') and ('avito.ru' in src or 'img.avito.ru' in src):
                    photo_urls.append(src)
            
            # Ищем изображения в background-image
            elements_with_bg = soup.find_all(style=re.compile(r'background-image'))
            for elem in elements_with_bg:
                style = elem.get('style', '')
                bg_match = re.search(r'background-image:\s*url\(["\']?([^"\')\s]+)["\']?\)', style)
                if bg_match:
                    bg_url = bg_match.group(1)
                    if bg_url.startswith('http') and ('avito.ru' in bg_url or 'img.avito.ru' in bg_url):
                        photo_urls.append(bg_url)
            
            # Убираем дубликаты, сохраняя порядок
            seen = set()
            unique_photos = []
            for url in photo_urls:
                if url not in seen:
                    seen.add(url)
                    unique_photos.append(url)
            
            return unique_photos
            
        except Exception as e:
            print(f"Ошибка при извлечении фотографий Avito: {e}")
            return []
    
    async def extract_photo_urls_from_url(self, listing_url: str) -> list[str]:
        """Асинхронно получает URL и извлекает ссылки на фотографии"""
        try:
            # Определяем источник ссылки
            if self.is_avito_url(listing_url):
                print(f"📸 Извлекаем фотографии с Avito: {listing_url}")
                # Для Avito используем более надежный способ для асинхронных запросов
                response = await asyncio.get_event_loop().run_in_executor(None, lambda: requests.get(listing_url, headers=HEADERS))
                response.raise_for_status()
                soup = BeautifulSoup(response.text, 'html.parser')
                return self.extract_avito_photo_urls(soup)
            else:
                print(f"📸 Извлекаем фотографии с Cian: {listing_url}")
                # Для Cian используем существующую логику
                response = await asyncio.get_event_loop().run_in_executor(None, lambda: requests.get(listing_url, headers=HEADERS))
                response.raise_for_status()
                soup = BeautifulSoup(response.text, 'html.parser')
                return self.extract_photo_urls(soup)
        except Exception as e:
            print(f"Ошибка при получении URL {listing_url}: {e}")
            return []
    
    async def extract_listing_info(self, url: str) -> dict:
        """Извлекает основную информацию об объявлении"""
        try:
            # Определяем источник ссылки
            if self.is_avito_url(url):
                print(f"🏠 Извлекаем информацию об объявлении Avito: {url}")
                # Для Avito используем асинхронный парсинг
                listing_data = await self.parse_avito_listing(url)
                if not listing_data:
                    return {
                        'rooms': 'N/A',
                        'price': 'N/A',
                        'floor': 'N/A',
                        'total_area': 'N/A',
                        'kitchen_area': 'N/A',
                        'metro': 'N/A'
                    }
                
                # Преобразуем данные Avito в формат для отображения
                info = {
                    'rooms': listing_data.get('rooms', 'N/A'),
                    'price': listing_data.get('price', 'N/A'),
                    'floor': listing_data.get('floor', 'N/A'),
                    'total_area': listing_data.get('total_area', 'N/A'),
                    'kitchen_area': listing_data.get('kitchen_area', 'N/A'),
                    'metro': listing_data.get('metro_time', 'N/A'),
                    'photo_urls': listing_data.get('photo_urls', [])  # Добавляем фотографии
                }
                return info
            else:
                # Для Cian используем существующую логику
                print(f"🏠 Извлекаем информацию об объявлении Cian: {url}")
                response = requests.get(url, headers=HEADERS)
                response.raise_for_status()
                soup = BeautifulSoup(response.text, 'html.parser')
                
                # Парсим основную информацию
                listing_data = parse_listing(url, requests.Session())
                
                # Формируем структурированную информацию
                info = {
                    'rooms': listing_data.get('Комнат', 'N/A'),
                    'price': listing_data.get('Цена_raw', 'N/A'),
                    'floor': listing_data.get('Этаж', 'N/A'),
                    'total_area': listing_data.get('Общая площадь', 'N/A'),
                    'kitchen_area': listing_data.get('Площадь кухни', 'N/A'),
                    'metro': listing_data.get('Минут метро', 'N/A')
                }
                return info
        except Exception as e:
            print(f"Ошибка при извлечении информации об объявлении {url}: {str(e)}")
            return {
                'rooms': 'N/A',
                'price': 'N/A',
                'floor': 'N/A',
                'total_area': 'N/A',
                'kitchen_area': 'N/A',
                'metro': 'N/A'
            }
    
    async def generate_html_gallery(self, listing_urls: list[str], user_id: int, subtitle: str = None, listing_comments: list[str] = None) -> str:
        """Генерирует HTML галерею с внешними ссылками на фотографии"""
        html_parts = []
        
        html_parts.append("""
        <!DOCTYPE html>
        <html lang="ru">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Подбор недвижимости</title>
            <style>
                body { font-family: Arial, sans-serif; margin: 20px; background-color: #f5f5f5; }
                .listing { 
                    background: white; 
                    margin: 20px 0; 
                    padding: 20px; 
                    border-radius: 10px; 
                    box-shadow: 0 2px 10px rgba(0,0,0,0.1);
                    display: flex;
                    gap: 30px;
                    align-items: flex-start;
                }
                
                .listing-info {
                    flex: 1;
                    min-width: 0;
                }
                
                .listing-photos {
                    flex: 1;
                    min-width: 0;
                }
                .listing h3 { color: #333; margin-top: 0; margin-bottom: 15px; }
                .listing p { margin: 8px 0; color: #555; }
                .listing strong { color: #333; }
                .main-title { color: #333; margin-bottom: 10px; }
                .subtitle { color: #666; font-size: 18px; margin-bottom: 30px; font-style: italic; }
                .photo-grid { 
                    display: grid; 
                    grid-template-columns: repeat(auto-fill, minmax(180px, 1fr)); 
                    gap: 8px; 
                    margin: 15px 0; 
                    max-height: 600px; 
                    overflow-y: auto; 
                    padding: 10px;
                    border: 1px solid #eee;
                    border-radius: 8px;
                }
                .photo-item { position: relative; }
                .photo-item img { 
                    width: 100%; 
                    height: 140px; 
                    object-fit: cover; 
                    border-radius: 5px; 
                    border: 2px solid transparent;
                    transition: border-color 0.2s;
                }
                .photo-item img:hover { 
                    border-color: #0066cc;
                    transition: border-color 0.2s ease;
                }
                
                /* Стили для модального окна */
                .modal {
                    display: none;
                    position: fixed;
                    z-index: 1000;
                    left: 0;
                    top: 0;
                    width: 100%;
                    height: 100%;
                    background-color: rgba(0,0,0,0.9);
                    overflow: auto;
                }
                
                .modal-content {
                    margin: auto;
                    display: block;
                    width: 90%;
                    max-width: 800px;
                    max-height: 90%;
                    position: relative;
                    top: 50%;
                    transform: translateY(-50%);
                }
                
                .modal-image {
                    width: 100%;
                    height: auto;
                    max-height: 90vh;
                    object-fit: contain;
                }
                
                .modal-caption {
                    margin: auto;
                    display: block;
                    width: 80%;
                    max-width: 700px;
                    text-align: center;
                    color: white;
                    padding: 20px 0;
                    font-size: 18px;
                }
                
                .close {
                    position: absolute;
                    top: 15px;
                    right: 35px;
                    color: #f1f1f1;
                    font-size: 40px;
                    font-weight: bold;
                    cursor: pointer;
                }
                
                .close:hover,
                .close:focus {
                    color: #bbb;
                    text-decoration: none;
                    cursor: pointer;
                }
                
                /* Кнопки навигации по фотографиям */
                .modal-nav {
                    position: absolute;
                    top: 50%;
                    transform: translateY(-50%);
                    background: rgba(0,0,0,0.5);
                    color: white;
                    border: none;
                    padding: 15px 12px;
                    cursor: pointer;
                    font-size: 24px;
                    border-radius: 50%;
                    width: 50px;
                    height: 50px;
                    display: flex;
                    align-items: center;
                    justify-content: center;
                    transition: background-color 0.3s;
                }
                
                .modal-nav:hover {
                    background: rgba(0,0,0,0.8);
                }
                
                .modal-nav.prev {
                    left: 20px;
                }
                
                .modal-nav.next {
                    right: 20px;
                }
                
                .modal-nav:disabled {
                    opacity: 0.3;
                    cursor: not-allowed;
                }
                
                /* Счетчик фотографий */
                .modal-counter {
                    position: absolute;
                    bottom: 20px;
                    left: 50%;
                    transform: translateX(-50%);
                    color: white;
                    background: rgba(0,0,0,0.7);
                    padding: 8px 16px;
                    border-radius: 20px;
                    font-size: 14px;
                }
                
                .no-photos { color: #666; font-style: italic; }
                
                /* Мобильная адаптация */
                @media (max-width: 768px) {
                    body { margin: 10px; }
                    .listing { 
                        padding: 15px; 
                        margin: 15px 0; 
                        flex-direction: column;
                        gap: 20px;
                        width: 100%;
                        box-sizing: border-box;
                    }
                    .listing-info, .listing-photos {
                        width: 100%;
                        min-width: 100%;
                    }
                    .photo-grid { 
                        grid-template-columns: repeat(auto-fill, minmax(150px, 1fr)); 
                        gap: 6px; 
                        padding: 8px;
                    }
                    .photo-item img { 
                        height: 120px; 
                    }
                    .main-title { font-size: 24px; }
                    .subtitle { font-size: 16px; }
                    
                    /* Мобильная адаптация для кнопок навигации */
                    .modal-nav {
                        width: 40px;
                        height: 40px;
                        font-size: 20px;
                        padding: 10px 8px;
                    }
                    
                    .modal-nav.prev {
                        left: 10px;
                    }
                    
                    .modal-nav.next {
                        right: 10px;
                    }
                    
                    .modal-counter {
                        bottom: 10px;
                        font-size: 12px;
                        padding: 6px 12px;
                    }
                }
            </style>
        </head>
        <body>
            <h1 class="main-title">🏠 Подбор недвижимости</h1>
            
            <!-- Модальное окно для фотографий -->
            <div id="photoModal" class="modal">
                <span class="close" onclick="closeModal()">&times;</span>
                <button class="modal-nav prev" id="prevBtn" onclick="showPrevPhoto()" title="Предыдущее фото (←)">‹</button>
                <button class="modal-nav next" id="nextBtn" onclick="showNextPhoto()" title="Следующее фото (→)">›</button>
                <img class="modal-content modal-image" id="modalImage">
                <div class="modal-caption" id="modalCaption"></div>
                <div class="modal-counter" id="modalCounter"></div>
            </div>
            
            <script>
                // Проверяем загрузку скрипта и элементов
                console.log('Скрипт загружен');
                
                // Глобальные переменные для навигации по фотографиям
                let currentPhotos = [];
                let currentPhotoIndex = 0;
                let currentListingIndex = 0;
                
                // Функция открытия модального окна
                function openModal(imageSrc, caption, photos, photoIndex, listingIndex) {
                    currentPhotos = photos || [];
                    currentPhotoIndex = photoIndex || 0;
                    currentListingIndex = listingIndex || 0;
                    
                    document.getElementById('photoModal').style.display = 'block';
                    document.getElementById('modalImage').src = imageSrc;
                    document.getElementById('modalCaption').innerHTML = caption;
                    
                    updateModalNavigation();
                }
                
                // Функция закрытия модального окна
                function closeModal() {
                    document.getElementById('photoModal').style.display = 'none';
                    currentPhotos = [];
                    currentPhotoIndex = 0;
                }
                
                // Функция показа предыдущего фото
                function showPrevPhoto() {
                    if (currentPhotos.length > 1) {
                        currentPhotoIndex = (currentPhotoIndex - 1 + currentPhotos.length) % currentPhotos.length;
                        const photo = currentPhotos[currentPhotoIndex];
                        document.getElementById('modalImage').src = photo.src;
                        document.getElementById('modalCaption').innerHTML = photo.caption;
                        updateModalNavigation();
                    }
                }
                
                // Функция показа следующего фото
                function showNextPhoto() {
                    if (currentPhotos.length > 1) {
                        currentPhotoIndex = (currentPhotoIndex + 1) % currentPhotos.length;
                        const photo = currentPhotos[currentPhotoIndex];
                        document.getElementById('modalImage').src = photo.src;
                        document.getElementById('modalCaption').innerHTML = photo.caption;
                        updateModalNavigation();
                    }
                }
                
                // Обновление навигации и счетчика
                function updateModalNavigation() {
                    const prevBtn = document.getElementById('prevBtn');
                    const nextBtn = document.getElementById('nextBtn');
                    const counter = document.getElementById('modalCounter');
                    
                    if (currentPhotos.length <= 1) {
                        prevBtn.style.display = 'none';
                        nextBtn.style.display = 'none';
                        counter.style.display = 'none';
                    } else {
                        prevBtn.style.display = 'flex';
                        nextBtn.style.display = 'flex';
                        counter.style.display = 'block';
                        
                        prevBtn.disabled = false;
                        nextBtn.disabled = false;
                        
                        counter.innerHTML = `${{currentPhotoIndex + 1}} из ${{currentPhotos.length}}`;
                    }
                }
                
                // Закрытие модального окна при клике вне изображения
                window.onclick = function(event) {
                    const modal = document.getElementById('photoModal');
                    if (event.target == modal) {
                        closeModal();
                    }
                }
                
                // Обработка клавиш для навигации
                document.addEventListener('keydown', function(event) {
                    if (event.key === 'Escape') {
                        closeModal();
                    } else if (event.key === 'ArrowLeft') {
                        showPrevPhoto();
                    } else if (event.key === 'ArrowRight') {
                        showNextPhoto();
                    }
                });
            </script>
        """)
        
        if subtitle:
            html_parts.append(f'<h2 class="subtitle">{subtitle}</h2>')
        
        html_parts.append("")
        
        # Собираем данные для сохранения в БД
        db_listings = []
        
        for i, listing_url in enumerate(listing_urls, 1):
            try:
                # Парсим объявление в зависимости от источника
                if self.is_avito_url(listing_url):
                    # Для Avito используем асинхронный парсинг
                    listing_data = await self.parse_avito_listing(listing_url)
                    
                    if not listing_data:
                        html_parts.append(f"""
                        <div class="listing">
                            <h3>Вариант #{i}</h3>
                            <p style="color: red;">Ошибка при парсинге Avito</p>
                        </div>
                        """)
                        continue
                    
                    # Сохраняем данные для БД
                    db_listings.append(listing_data)
                    
                    # Преобразуем данные Avito в формат для отображения
                    listing_data_display = {
                        'Комнат': listing_data.get('rooms', 'N/A'),
                        'Цена_raw': listing_data.get('price', 'N/A'),
                        'floor': listing_data.get('floor', 'N/A'),
                        'total_floors': listing_data.get('total_floors', 'N/A'),
                        'Общая площадь': listing_data.get('total_area', 'N/A'),
                        'Площадь кухни': listing_data.get('kitchen_area', 'N/A'),
                        'Минут метро': listing_data.get('metro_time', 'N/A')
                    }
                    
                    # Для Avito фотографии не извлекаем (требует отдельной логики)
                    photo_urls = []
                else:
                    # Для Cian используем существующую логику
                    listing_data = parse_listing(listing_url, requests.Session())
                    listing_data_display = listing_data
                    
                    # Извлекаем фотографии
                    soup = BeautifulSoup(requests.get(listing_url, headers=HEADERS).text, 'html.parser')
                    photo_urls = self.extract_photo_urls(soup)
                    
                    # Сохраняем данные Cian для БД
                    cian_data = {
                        'url': listing_url,
                        'source': 4,  # Cian
                        'rooms': listing_data.get('Комнат', 'N/A'),
                        'price': listing_data.get('Цена_raw', 'N/A'),
                        'floor': listing_data.get('Этаж', 'N/A'),
                        'total_floors': listing_data.get('total_floors', 'N/A'),
                        'total_area': listing_data.get('Общая площадь', 'N/A'),
                        'kitchen_area': listing_data.get('Площадь кухни', 'N/A'),
                        'metro_time': listing_data.get('Минут метро', 'N/A'),
                        'photo_urls': photo_urls if photo_urls else []
                    }
                    db_listings.append(cian_data)
                
                html_parts.append(f"""
                <div class="listing">
                    <h3>Вариант #{i}</h3>
                """)
                
                # Добавляем комментарий к объявлению, если есть
                if listing_comments and i <= len(listing_comments) and listing_comments[i-1]:
                    html_parts.append(f'<p class="listing-comment">{listing_comments[i-1]}</p>')
                
                html_parts.append("")
                
                # Добавляем основную информацию
                html_parts.append(f"<p><strong>Комнат:</strong> {listing_data_display.get('Комнат', 'N/A')}</p>")
                html_parts.append(f"<p><strong>Цена:</strong> {listing_data_display.get('Цена_raw', 'N/A')}</p>")
                
                # Отображаем этаж правильно для Avito и Cian
                # Для Cian используем поле 'Этаж', для Avito - 'floor'
                floor_value = listing_data_display.get('floor', listing_data_display.get('Этаж', 'N/A'))
                total_floors = listing_data_display.get('total_floors')
                if total_floors and total_floors != 'N/A':
                    html_parts.append(f"<p><strong>Этаж:</strong> {floor_value}/{total_floors}</p>")
                else:
                    html_parts.append(f"<p><strong>Этаж:</strong> {floor_value}</p>")
                
                html_parts.append(f"<p><strong>Общая площадь:</strong> {listing_data_display.get('Общая площадь', 'N/A')} м²</p>")
                html_parts.append(f"<p><strong>Кухня:</strong> {listing_data_display.get('Площадь кухни', 'N/A')} м²</p>")

                
                # Создаем массив фотографий для модального окна
                modal_photos = []
                if photo_urls:
                    for idx, photo_url in enumerate(photo_urls):
                        modal_photos.append({
                            'src': photo_url,
                            'caption': f"Фото {idx + 1}"
                        })
                
                # Добавляем фотографии (только для Cian)
                if photo_urls:
                    # Генерируем сетку фотографий (все фото без ограничений)
                    html_parts.append(f'<div class="photo-grid">')
                    for j, photo_url in enumerate(photo_urls):
                        html_parts.append(f"""
                        <div class="photo-item">
                            <img src="{photo_url}" alt="Фото {j+1}" 
                                 onerror="this.style.display='none'; this.nextElementSibling.style.display='block';"
                                 loading="lazy"
                                 onclick="openModal('{photo_url}', 'Фото {j+1}', {json.dumps(modal_photos)}, {j}, {i})"
                                 style="cursor: pointer;">
                            <div class="photo-fallback" style="display: none; background: #f0f0f0; border: 1px dashed #ccc; border-radius: 5px; padding: 20px; text-align: center; color: #666;">
                                <div>📷 Фото {j+1}</div>
                                <div style="font-size: 12px; margin-top: 5px;">
                                    <a href="{photo_url}" target="_blank" style="color: #0066cc;">Открыть фото</a>
                                </div>
                            </div>
                        </div>
                        """)
                    html_parts.append('</div>')
                else:
                    if self.is_avito_url(listing_url):
                        html_parts.append('<p class="no-photos">📷 Фотографии Avito (требуют отдельной обработки)</p>')
                    else:
                        html_parts.append('<p class="no-photos">📷 Фотографии не найдены</p>')
                
                html_parts.append('</div>')
                
            except Exception as e:
                html_parts.append(f"""
                <div class="listing">
                    <h3>Вариант #{i}</h3>
                    <p style="color: red;">Ошибка при парсинге: {str(e)}</p>
                </div>
                """)
        
        html_parts.append("""
        </body>
        </html>
        """)
        
        # Проверяем размер HTML
        html_content = ''.join(html_parts)
        html_size_mb = len(html_content.encode('utf-8')) / 1024 / 1024
        print(f"📊 Размер HTML: {html_size_mb:.1f}MB")
        
        if html_size_mb > 10:  # Предупреждение при больших файлах
            print(f"⚠️  ВНИМАНИЕ: HTML файл очень большой ({html_size_mb:.1f}MB)!")
            print(f"⚠️  Это может вызвать проблемы при передаче данных.")
            
            # Если файл слишком большой, создаем упрощенную версию
            if html_size_mb > 15:  # Критически большой файл
                print(f"🚨 HTML файл критически большой ({html_size_mb:.1f}MB), создаем упрощенную версию...")
                
                # Создаем упрощенную версию без фотографий
                simplified_html = f"""
                <!DOCTYPE html>
                <html lang="ru">
                <head>
                    <meta charset="UTF-8">
                    <meta name="viewport" content="width=device-width, initial-scale=1.0">
                    <title>Подбор недвижимости (упрощенная версия)</title>
                    <style>
                        body {{ font-family: Arial, sans-serif; margin: 20px; background-color: #f5f5f5; }}
                        .listing {{ background: white; margin: 20px 0; padding: 20px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
                        .warning {{ background: #fff3cd; border: 1px solid #ffeaa7; padding: 15px; border-radius: 5px; margin: 20px 0; }}
                    </style>
                </head>
                <body>
                    <h1 class="main-title">Подбор недвижимости</h1>
                    <div class="warning">
                        <strong>⚠️ Внимание:</strong> Исходный файл был слишком большим ({html_size_mb:.1f}MB), 
                        поэтому отображается упрощенная версия без фотографий. 
                        Для просмотра полной версии с фотографиями попробуйте обработать меньше объявлений за раз.
                    </div>
                """
                
                # Добавляем только текстовую информацию об объявлениях
                for i, listing_url in enumerate(listing_urls, 1):
                    simplified_html += f"""
                    <div class="listing">
                        <h3>Вариант #{i}</h3>
                        <p><strong>Ссылка:</strong> <a href="{listing_url}" target="_blank">{listing_url}</a></p>
                        <p><strong>Статус:</strong> Фотографии недоступны (файл слишком большой)</p>
                    </div>
                    """
                
                simplified_html += """
                </body>
                </html>
                """
                
                simplified_size_mb = len(simplified_html.encode('utf-8')) / 1024 / 1024
                print(f"✅ Создана упрощенная версия размером {simplified_size_mb:.1f}MB")
                
                return simplified_html
        
        # Сохраняем все объявления в БД
        if db_listings:
            try:
                print(f"💾 Сохраняем {len(db_listings)} объявлений в БД...")
                await save_listings(db_listings, user_id)
                print(f"✅ Объявления успешно сохранены в БД")
            except Exception as e:
                print(f"❌ Ошибка сохранения в БД: {e}")
        else:
            print(f"⚠️ Нет данных для сохранения в БД")
        
        return html_content
    
    async def generate_html_gallery_embedded(self, listing_urls: list[str], user_id: int, subtitle: str = None, remove_watermarks: bool = False, max_photos_per_listing: int = None, listing_comments: list[str] = None, pre_parsed_data: dict = None) -> tuple[str, list[dict]]:
        """Генерирует HTML галерею с встроенными Base64 изображениями и возвращает статистику по фото"""
        html_content = f"""
        <!DOCTYPE html>
        <html lang="ru">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Подбор недвижимости</title>
            <style>
                body {{ 
                    font-family: Arial, sans-serif; 
                    margin: 20px; 
                    background-color: #f5f5f5; 
                }}
                .listing {{ 
                    background: white; 
                    margin: 20px 0; 
                    padding: 20px; 
                    border-radius: 10px; 
                    box-shadow: 0 2px 10px rgba(0,0,0,0.1);
                }}
                
                .listing-header {{
                    margin-bottom: 20px;
                }}
                
                .listing-info {{
                    flex: 1;
                    min-width: 0;
                }}
                

                
                .listing-photos {{
                    width: 100%;
                }}
                .listing h3 {{ 
                    color: #333; 
                    margin-top: 0; 
                    margin-bottom: 15px; 
                }}
                .listing p {{ 
                    margin: 8px 0; 
                    color: #555; 
                }}
                .listing strong {{ 
                    color: #333; 
                }}
                
                /* Стили для комментариев к объявлениям */
                .listing-comment {{
                    background: #f8f9fa;
                    padding: 12px;
                    margin: 10px 0;
                    border-left: 4px solid #0066cc;
                    border-radius: 5px;
                    font-style: italic;
                    color: #555;
                    font-size: 14px;
                    line-height: 1.4;
                }}
                .main-title {{ 
                    color: #333; 
                    margin-bottom: 10px; 
                }}
                .subtitle {{ 
                    color: #666; 
                    font-size: 18px; 
                    margin-bottom: 30px; 
                    font-style: italic; 
                }}
                .photo-grid {{ 
                    display: grid; 
                    grid-template-columns: repeat(auto-fill, minmax(180px, 1fr)); 
                    gap: 8px; 
                    margin: 15px 0; 
                    max-height: 600px; 
                    overflow-y: auto; 
                    padding: 10px;
                    border: 1px solid #eee;
                    border-radius: 8px;
                }}
                .photo-item {{ 
                    position: relative; 
                }}
                .photo-item img {{ 
                    width: 100%; 
                    height: 140px; 
                    object-fit: cover; 
                    border-radius: 5px; 
                    border: 2px solid transparent;
                    transition: border-color 0.2s;
                    background: #f8f9fa;
                }}
                

                
                .photo-item img:hover {{
                    border-color: #0066cc;
                    cursor: pointer;
                    transform: scale(1.02);
                    transition: transform 0.2s ease;
                }}
                
                /* Стили для модального окна */
                .modal {{
                    display: none;
                    position: fixed;
                    z-index: 1000;
                    left: 0;
                    top: 0;
                    width: 100%;
                    height: 100%;
                    background-color: rgba(0,0,0,0.9);
                    overflow: auto;
                }}
                
                .modal-content {{
                    margin: auto;
                    display: block;
                    width: 90%;
                    max-width: 800px;
                    max-height: 90%;
                    position: relative;
                    top: 50%;
                    transform: translateY(-50%);
                }}
                
                .modal-image {{
                    width: 100%;
                    height: auto;
                    max-height: 90vh;
                    object-fit: contain;
                }}
                
                .modal-caption {{
                    margin: auto;
                    display: block;
                    width: 80%;
                    max-width: 700px;
                    text-align: center;
                    color: white;
                    padding: 20px 0;
                    font-size: 18px;
                }}
                
                .close {{
                    position: absolute;
                    top: 15px;
                    right: 35px;
                    color: #f1f1f1;
                    font-size: 40px;
                    font-weight: bold;
                    cursor: pointer;
                }}
                
                .close:hover,
                .close:focus {{
                    color: #bbb;
                    text-decoration: none;
                    cursor: pointer;
                }}
                
                .photo-fallback {{ 
                    width: 100%; 
                    height: 140px; 
                    display: flex; 
                    flex-direction: column; 
                    justify-content: center; 
                    align-items: center;
                }}
                .no-photos {{ 
                    color: #666; 
                    font-style: italic; 
                }}
                .photo-info {{ 
                    background: #f8f9fa; 
                    padding: 15px; 
                    border-radius: 8px; 
                    margin-top: 15px; 
                    font-size: 14px;
                    border-left: 4px solid #0066cc;
                }}
                .photo-info strong {{ 
                    color: #333; 
                }}
                .photo-info small {{ 
                    line-height: 1.4; 
                }}
                
                /* Мобильная адаптация */
                @media (max-width: 768px) {{
                    body {{ margin: 10px; }}
                    .listing {{ 
                        padding: 15px; 
                        margin: 15px 0; 
                        width: 100%;
                        box-sizing: border-box;
                    }}

                    .listing-info {{
                        width: 100%;
                        min-width: 100%;
                    }}
                    .listing-photos {{
                        width: 100%;
                    }}
                    .photo-grid {{ 
                        grid-template-columns: repeat(auto-fill, minmax(150px, 1fr)); 
                        gap: 6px; 
                        padding: 8px;
                    }}
                    .photo-item img, .photo-fallback {{ 
                        height: 120px; 
                    }}
                    

                    
                    .main-title {{ font-size: 24px; }}
                    .subtitle {{ font-size: 16px; }}
                    
                    /* Мобильная адаптация для кнопок навигации */
                    .modal-nav {{
                        width: 40px;
                        height: 40px;
                        font-size: 20px;
                        padding: 10px 8px;
                    }}
                    
                    .modal-nav.prev {{
                        left: 10px;
                    }}
                    
                    .modal-nav.next {{
                        right: 10px;
                    }}
                    
                    .modal-counter {{
                        bottom: 10px;
                        font-size: 12px;
                        padding: 6px 12px;
                    }}
                }}
            </style>
        </head>
        <body>
            <h1 class="main-title">🏠 Подбор недвижимости</h1>
            
            <!-- Модальное окно для фотографий -->
            <div id="photoModal" class="modal">
                <span class="close" onclick="closeModal()">&times;</span>
                <button class="modal-nav prev" id="prevBtn" onclick="showPrevPhoto()" title="Предыдущее фото (←)">‹</button>
                <button class="modal-nav next" id="nextBtn" onclick="showNextPhoto()" title="Следующее фото (→)">›</button>
                <img class="modal-content modal-image" id="modalImage">
                <div class="modal-caption" id="modalCaption"></div>
                <div class="modal-counter" id="modalCounter"></div>
            </div>
            
            <script>
                // Проверяем, что модальное окно существует
                console.log('Скрипт загружен');
                console.log('Модальное окно:', document.getElementById('photoModal'));
                console.log('Кнопка prev:', document.getElementById('prevBtn'));
                console.log('Кнопка next:', document.getElementById('nextBtn'));
                
                // Глобальные переменные для навигации по фотографиям
                let currentPhotos = [];
                let currentPhotoIndex = 0;
                let currentListingIndex = 0;
                
                // Функция открытия модального окна
                function openModal(imageSrc, caption, photos, photoIndex, listingIndex) {{
                    console.log('openModal вызвана с параметрами:', imageSrc, caption, photos, photoIndex, listingIndex);
                    
                    currentPhotos = photos || [];
                    currentPhotoIndex = photoIndex || 0;
                    currentListingIndex = listingIndex || 0;
                    
                    console.log('Элементы модального окна:', {
                        modal: document.getElementById('photoModal'),
                        image: document.getElementById('modalImage'),
                        caption: document.getElementById('modalCaption')
                    });
                    
                    document.getElementById('photoModal').style.display = 'block';
                    document.getElementById('modalImage').src = imageSrc;
                    document.getElementById('modalCaption').innerHTML = caption;
                    
                    updateModalNavigation();
                    console.log('Модальное окно должно быть открыто');
                }}
                
                // Функция закрытия модального окна
                function closeModal() {{
                    document.getElementById('photoModal').style.display = 'none';
                    currentPhotos = [];
                    currentPhotoIndex = 0;
                }}
                
                // Функция показа предыдущего фото
                function showPrevPhoto() {{
                    if (currentPhotos.length > 1) {{
                        currentPhotoIndex = (currentPhotoIndex - 1 + currentPhotos.length) % currentPhotos.length;
                        const photo = currentPhotos[currentPhotoIndex];
                        document.getElementById('modalImage').src = photo.src;
                        document.getElementById('modalCaption').innerHTML = photo.caption;
                        updateModalNavigation();
                    }}
                }}
                
                // Функция показа следующего фото
                function showNextPhoto() {{
                    if (currentPhotos.length > 1) {{
                        currentPhotoIndex = (currentPhotoIndex + 1) % currentPhotos.length;
                        const photo = currentPhotos[currentPhotoIndex];
                        document.getElementById('modalImage').src = photo.src;
                        document.getElementById('modalCaption').innerHTML = photo.caption;
                        updateModalNavigation();
                    }}
                }}
                
                // Обновление навигации и счетчика
                function updateModalNavigation() {{
                    const prevBtn = document.getElementById('prevBtn');
                    const nextBtn = document.getElementById('nextBtn');
                    const counter = document.getElementById('modalCounter');
                    
                    if (currentPhotos.length <= 1) {{
                        prevBtn.style.display = 'none';
                        nextBtn.style.display = 'none';
                        counter.style.display = 'none';
                    }} else {{
                        prevBtn.style.display = 'flex';
                        nextBtn.style.display = 'flex';
                        counter.style.display = 'block';
                        
                        prevBtn.disabled = false;
                        nextBtn.disabled = false;
                        
                        counter.innerHTML = `${{currentPhotoIndex + 1}} из ${{currentPhotos.length}}`;
                    }}
                }}
                
                // Закрытие модального окна при клике вне изображения
                window.onclick = function(event) {{
                    const modal = document.getElementById('photoModal');
                    if (event.target == modal) {{
                        closeModal();
                    }}
                }}
                
                // Обработка клавиш для навигации
                document.addEventListener('keydown', function(event) {{
                    if (event.key === 'Escape') {{
                        closeModal();
                    }} else if (event.key === 'ArrowLeft') {{
                        showPrevPhoto();
                    }} else if (event.key === 'ArrowRight') {{
                        showNextPhoto();
                    }}
                }});
                
                // Проверяем элементы после загрузки страницы
                document.addEventListener('DOMContentLoaded', function() {{
                    console.log('DOM загружен, проверяем элементы:');
                    console.log('Модальное окно:', document.getElementById('photoModal'));
                    console.log('Кнопка prev:', document.getElementById('prevBtn'));
                    console.log('Кнопка next:', document.getElementById('nextBtn'));
                    console.log('Изображение:', document.getElementById('modalImage'));
                    console.log('Подпись:', document.getElementById('modalCaption'));
                    console.log('Счетчик:', document.getElementById('modalCounter'));
                }});
            </script>
        """
        
        if subtitle:
            html_content += f'<h2 class="subtitle">{subtitle}</h2>'
        
        html_content += """
        """
        
        # Собираем статистику по фото для каждого объявления
        photo_stats = []
        
        # Собираем данные для сохранения в БД
        db_listings = []
        
        for i, listing_url in enumerate(listing_urls, 1):
            try:
                print(f"🔍 Обрабатываю объявление {i}: {listing_url}")
                
                # Парсим объявление в зависимости от источника
                if self.is_avito_url(listing_url):
                    # Для Avito используем асинхронный парсинг
                    # Если это первый URL и у нас есть предварительно спарсенные данные, используем их
                    if i == 1 and pre_parsed_data:
                        listing_data = pre_parsed_data
                        print(f"🔄 Используем предварительно спарсенные данные для URL #{i}")
                    else:
                        listing_data = await self.parse_avito_listing(listing_url)
                    
                    if not listing_data:
                        html_content += f"""
                        <div class="listing">
                            <div class="listing-header">
                                <div class="listing-info">
                                    <h3>Вариант #{i}</h3>
                                    <p>Ошибка при парсинге Avito</p>
                                </div>
                            </div>
                            <div class="listing-photos">
                                <p>📷 Фотографии недоступны</p>
                            </div>
                        </div>
                        """
                        continue
                    
                    # Сохраняем данные для БД
                    db_listings.append(listing_data)
                    
                    # Преобразуем данные Avito в формат для отображения
                    listing_info = {
                        'rooms': listing_data.get('rooms', 'N/A'),
                        'price': listing_data.get('price', 'N/A'),
                        'floor': listing_data.get('floor', 'N/A'),
                        'total_floors': listing_data.get('total_floors', 'N/A'),
                        'total_area': listing_data.get('total_area', 'N/A'),
                        'kitchen_area': listing_data.get('kitchen_area', 'N/A'),
                        'metro': listing_data.get('metro_time', 'N/A')
                    }
                    
                    # Обрабатываем фотографии Avito
                    processed_photos = []
                    
                    if 'photo_urls' in listing_data and listing_data['photo_urls']:
                        photo_urls = listing_data['photo_urls']
                        print(f"📸 Найдено {len(photo_urls)} фотографий Avito")
                        
                        # Сначала пробуем обработать через photo_processor
                        if photo_urls:
                            try:
                                processed_photos = self.photo_processor.process_photos_for_embedded_html(
                                    photo_urls, remove_watermarks, max_photos=8
                                )
                                print(f"✅ Обработано {len(processed_photos)} фотографий Avito через photo_processor")
                            except Exception as e:
                                print(f"⚠️ Ошибка обработки фотографий Avito через photo_processor: {e}")
                                processed_photos = []
                        
                        # Если photo_processor не сработал, создаем простые ссылки
                        if not processed_photos and photo_urls:
                            print(f"🔄 Создаем простые ссылки для {len(photo_urls)} фотографий Avito")
                            for idx, url in enumerate(photo_urls):
                                processed_photos.append({
                                    'format': 'jpeg',
                                    'base64': None,
                                    'url': url,
                                    'is_external': True
                                })
                            print(f"✅ Создано {len(processed_photos)} простых ссылок")
                    else:
                        print(f"⚠️ Фотографии Avito не найдены в данных")
                else:
                    # Для Cian используем существующую логику
                    listing_info = await self.extract_listing_info(listing_url)
                    
                    # Извлекаем URL фотографий
                    photo_urls = await self.extract_photo_urls_from_url(listing_url)
                    
                    # Обрабатываем фотографии
                    if photo_urls:
                        processed_photos = self.photo_processor.process_photos_for_embedded_html(
                            photo_urls, remove_watermarks, max_photos=8
                        )
                    else:
                        processed_photos = []
                    
                    # Сохраняем данные Cian для БД
                    cian_data = {
                        'url': listing_url,
                        'source': 4,  # Cian
                        'rooms': listing_info.get('rooms', 'N/A'),
                        'price': listing_info.get('price', 'N/A'),
                        'floor': listing_info.get('floor', listing_info.get('Этаж', 'N/A')),
                        'total_floors': listing_info.get('total_floors', 'N/A'),
                        'total_area': listing_info.get('total_area', 'N/A'),
                        'kitchen_area': listing_info.get('kitchen_area', 'N/A'),
                        'metro_time': listing_info.get('metro', 'N/A'),
                        'photo_urls': photo_urls if photo_urls else []
                    }
                    db_listings.append(cian_data)
                
                html_content += f"""
                <div class="listing">
                    <div class="listing-header">
                        <div class="listing-info">
                            <h3>Вариант #{i}</h3>
                """
                
                # Добавляем комментарий к объявлению, если есть
                if listing_comments and i <= len(listing_comments) and listing_comments[i-1]:
                    html_content += f'<p class="listing-comment">{listing_comments[i-1]}</p>'
                
                html_content += f"""
                            <p><strong>Комнат:</strong> {listing_info.get('rooms', 'N/A')}</p>
                            <p><strong>Цена:</strong> {listing_info.get('price', 'N/A')}</p>
                            <p><strong>Этаж:</strong> {listing_info.get('floor', 'N/A')}{'/' + str(listing_info.get('total_floors')) if listing_info.get('total_floors') and listing_info.get('total_floors') != 'N/A' else ''}</p>
                            <p><strong>Общая площадь:</strong> {listing_info.get('total_area', 'N/A')} м²</p>
                            <p><strong>Кухня:</strong> {listing_info.get('kitchen_area', 'N/A')} м²</p>
                        </div>
                    </div>
                    
                    <div class="listing-photos">
                """
                
                # Создаем массив фотографий для модального окна
                modal_photos = []
                
                if processed_photos:
                    for idx, photo in enumerate(processed_photos):
                        if photo and 'base64' in photo:
                            # Base64 изображение - проверяем размер
                            base64_size_mb = len(photo['base64']) / 1024 / 1024
                            if base64_size_mb > 2:  # Пропускаем слишком большие изображения
                                print(f"⚠️  Пропускаем фото {idx+1} для modal_photos - слишком большое ({base64_size_mb:.1f}MB)")
                                continue
                                
                            modal_photos.append({
                                'src': f"data:image/{photo['format']};base64,{photo['base64']}",
                                'caption': f"Фото {idx + 1}" if idx > 0 else "Главное фото"
                            })
                        elif photo and 'url' in photo:
                            # Внешняя ссылка
                            modal_photos.append({
                                'src': photo['url'],
                                'caption': f"Фото {idx + 1}" if idx > 0 else "Главное фото"
                            })
                
                                # Добавляем сетку со всеми фотографиями
                if processed_photos and len(processed_photos) > 0:
                    html_content += '<div class="photo-grid">'
                    for j, photo_data in enumerate(processed_photos):
                        if photo_data and 'base64' in photo_data:
                            # Base64 изображение - проверяем размер
                            base64_size_mb = len(photo_data['base64']) / 1024 / 1024
                            if base64_size_mb > 2:  # Пропускаем слишком большие изображения
                                print(f"⚠️  Пропускаем фото {j+1} - слишком большое ({base64_size_mb:.1f}MB)")
                                continue
                                
                            photo_caption = "Главное фото" if j == 0 else f"Фото {j + 1}"
                            html_content += f"""
                            <div class="photo-item">
                                <img src="data:image/{photo_data['format']};base64,{photo_data['base64']}" 
                                     alt="{photo_caption}" 
                                     loading="lazy"
                                     onclick="console.log('Клик по фото {j}'); openModal('data:image/{photo_data['format']};base64,{photo_data['base64']}', '{photo_caption}', {json.dumps(modal_photos)}, {j}, {i})"
                                     style="cursor: pointer;">
                            </div>
                            """
                        elif photo_data and 'url' in photo_data:
                            # Внешняя ссылка
                            photo_caption = "Главное фото" if j == 0 else f"Фото {j + 1}"
                            html_content += f"""
                            <div class="photo-item">
                                <img src="{photo_data['url']}" 
                                     alt="{photo_caption}" 
                                     loading="lazy"
                                     onclick="console.log('Клик по фото {j}'); openModal('{photo_data['url']}', '{photo_caption}', {json.dumps(modal_photos)}, {j}, {i})"
                                     style="cursor: pointer;">
                            </div>
                            """
                    
                    html_content += '</div>'
                else:
                    html_content += '<div class="photo-grid">'
                
                # Добавляем сообщение, если фотографий нет
                if not processed_photos:
                    if self.is_avito_url(listing_url):
                        html_content += '<p>📷 Фотографии Avito не найдены</p>'
                    else:
                        html_content += '<p>📷 Фотографии не найдены</p>'
                else:
                    html_content += f'<p>📷 Найдено фотографий: {len(processed_photos)}</p>'
                
                # Сохраняем статистику по фото
                photo_stats.append({
                    'listing_number': i,
                    'photo_count': len(processed_photos) if processed_photos else 0,
                    'url': listing_url
                })
                
                html_content += f"""
                    </div>
                </div>
                """
            except Exception as e:
                # Сохраняем статистику для объявления с ошибкой
                photo_stats.append({
                    'listing_number': i,
                    'photo_count': 0,
                    'url': listing_url,
                    'error': str(e)
                })
                
                html_content += f"""
                <div class="listing">
                    <div class="listing-header">
                        <div class="listing-info">
                            <h3>Вариант #{i}</h3>
                            <p>Ошибка при обработке: {str(e)}</p>
                        </div>
                    </div>
                    <div class="listing-photos">
                        <p>📷 Фотографии недоступны</p>
                    </div>
                </div>
                """
        
        html_content += """
        </body>
        </html>
        """
        
        # Проверяем размер HTML
        html_size_mb = len(html_content.encode('utf-8')) / 1024 / 1024
        print(f"📊 Размер HTML: {html_size_mb:.1f}MB")
        
        if html_size_mb > 10:  # Предупреждение при больших файлах
            print(f"⚠️  ВНИМАНИЕ: HTML файл очень большой ({html_size_mb:.1f}MB)!")
            print(f"⚠️  Это может вызвать проблемы при передаче данных.")
            
            # Если файл слишком большой, создаем упрощенную версию
            if html_size_mb > 15:  # Критически большой файл
                print(f"🚨 HTML файл критически большой ({html_size_mb:.1f}MB), создаем упрощенную версию...")
                
                # Создаем упрощенную версию без фотографий
                simplified_html = f"""
                <!DOCTYPE html>
                <html lang="ru">
                <head>
                    <meta charset="UTF-8">
                    <meta name="viewport" content="width=device-width, initial-scale=1.0">
                    <title>Подбор недвижимости (упрощенная версия)</title>
                    <style>
                        body {{ font-family: Arial, sans-serif; margin: 20px; background-color: #f5f5f5; }}
                        .listing {{ background: white; margin: 20px 0; padding: 20px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
                        .warning {{ background: #fff3cd; border: 1px solid #ffeaa7; padding: 15px; border-radius: 5px; margin: 20px 0; }}
                    </style>
                </head>
                <body>
                    <h1 class="main-title">Подбор недвижимости</h1>
                    <div class="warning">
                        <strong>⚠️ Внимание:</strong> Исходный файл был слишком большим ({html_size_mb:.1f}MB), 
                        поэтому отображается упрощенная версия без фотографий. 
                        Для просмотра полной версии с фотографиями попробуйте обработать меньше объявлений за раз.
                    </div>
                """
                
                # Добавляем только текстовую информацию об объявлениях
                for i, listing_url in enumerate(listing_urls, 1):
                    simplified_html += f"""
                    <div class="listing">
                        <h3>Вариант #{i}</h3>
                        <p><strong>Ссылка:</strong> <a href="{listing_url}" target="_blank">{listing_url}</a></p>
                        <p><strong>Статус:</strong> Фотографии недоступны (файл слишком большой)</p>
                    </div>
                    """
                
                simplified_html += """
                </body>
                </html>
                """
                
                simplified_size_mb = len(simplified_html.encode('utf-8')) / 1024 / 1024
                print(f"✅ Создана упрощенная версия размером {simplified_size_mb:.1f}MB")
                
                return simplified_html, []  # Возвращаем пустую статистику для упрощенной версии
        
        # Сохраняем все объявления в БД
        if db_listings:
            try:
                print(f"💾 Сохраняем {len(db_listings)} объявлений в БД...")
                await save_listings(db_listings, user_id)
                print(f"✅ Объявления успешно сохранены в БД")
            except Exception as e:
                print(f"❌ Ошибка сохранения в БД: {e}")
        else:
            print(f"⚠️ Нет данных для сохранения в БД")
        
        return html_content, photo_stats

    async def parse_listings_batch(self, listing_urls: list[str]) -> list[dict]:
        """Универсальный метод для парсинга списка объявлений с Cian и Avito"""
        parsed_listings = []
        
        for i, url in enumerate(listing_urls, 1):
            try:
                print(f"🔄 Парсим объявление {i}/{len(listing_urls)}: {url}")
                
                if self.is_avito_url(url):
                    print(f"🏠 Источник: Avito")
                    listing_data = await self.parse_avito_listing(url)
                    if listing_data:
                        # Добавляем источник
                        listing_data['source'] = 1  # Avito
                        parsed_listings.append(listing_data)
                        print(f"✅ Объявление Avito успешно спарсено")
                    else:
                        print(f"❌ Не удалось спарсить объявление Avito")
                elif self.is_cian_url(url):
                    print(f"🏠 Источник: Cian")
                    # Для Cian используем существующую логику
                    session = requests.Session()
                    listing_data = parse_listing(url, session)
                    if listing_data:
                        # Добавляем источник
                        listing_data['source'] = 4  # Cian
                        parsed_listings.append(listing_data)
                        print(f"✅ Объявление Cian успешно спарсено")
                    else:
                        print(f"❌ Не удалось спарсить объявление Cian")
                else:
                    print(f"⚠️ Неизвестный источник ссылки: {url}")
                
            except Exception as e:
                print(f"❌ Ошибка при парсинге {url}: {e}")
                continue
        
        print(f"📊 Всего успешно спарсено: {len(parsed_listings)} из {len(listing_urls)}")
        return parsed_listings
    
    def cleanup(self):
        """Корректно закрывает ресурсы"""
        try:
            # PhotoProcessor не имеет метода cleanup, пропускаем
            pass
        except Exception as e:
            print(f"⚠️ Ошибка при очистке ресурсов: {e}")
    
    def __del__(self):
        """Деструктор для автоматической очистки при удалении объекта"""
        self.cleanup()

def extract_number(text: str):
    if not text or text == '—':
        return None
    cleaned = re.sub(r"[^\d.,]", "", text)
    cleaned = cleaned.replace('\u00A0', '').replace(' ', '').replace(',', '.')
    try:
        return float(cleaned) if '.' in cleaned else int(cleaned)
    except ValueError:
        return None

async def export_listings_to_excel(listing_urls: list[str], user_id: int, output_path: str = None) -> tuple[BytesIO, int]:
    """
    Парсит список объявлений, сохраняет их в БД и возвращает Excel-файл и request_id.
    Поддерживает как Cian, так и Avito.
    :param listing_urls: список URL объявлений
    :param user_id: ID пользователя для сохранения в БД
    :param output_path: опциональный путь для сохранения файла на диск
    :return: tuple (BytesIO с данными файла, request_id)
    """
    sess = requests.Session()
    rows = []
    
    # Создаем экземпляр процессора для определения источника ссылок
    processor = ListingsProcessor()
    
    for url in listing_urls:
        try:
            if processor.is_avito_url(url):
                print(f"🏠 Парсим объявление Avito: {url}")
                # Для Avito используем асинхронный парсинг
                avito_data = await processor.parse_avito_listing(url)
                if avito_data:
                    # Преобразуем данные Avito в формат для Excel
                    excel_data = {
                        'URL': url,
                        'Комнат': avito_data.get('rooms', 'N/A'),
                        'Цена_raw': avito_data.get('price', 'N/A'),
                        'Этаж': f"{avito_data.get('floor', 'N/A')}/{avito_data.get('total_floors')}" if avito_data.get('total_floors') and avito_data.get('total_floors') != 'N/A' else f"{avito_data.get('floor', 'N/A')}",
                        'Общая площадь': avito_data.get('total_area', 'N/A'),
                        'Жилая площадь': avito_data.get('living_area', 'N/A'),
                        'Площадь кухни': avito_data.get('kitchen_area', 'N/A'),
                        'Санузел': avito_data.get('bathroom', 'N/A'),
                        'Балкон/лоджия': avito_data.get('balcony', 'N/A'),
                        'Вид из окон': avito_data.get('windows', 'N/A'),
                        'Ремонт': avito_data.get('renovation', 'N/A'),
                        'Год постройки': avito_data.get('construction_year', 'N/A'),
                        'Строительная серия': 'N/A',  # Пусто в Avito
                        'Тип дома': avito_data.get('house_type', 'N/A'),
                        'Тип перекрытий': 'N/A',  # Пусто в Avito
                        'Пассажирских лифтов': avito_data.get('passenger_elevator', 'N/A'),
                        'Грузовых лифтов': avito_data.get('cargo_elevator', 'N/A'),
                        'Парковка': avito_data.get('parking', 'N/A'),
                        'Газоснабжение': avito_data.get('gas_supply', 'N/A'),  # Берем из "В доме"
                        'Высота потолков': avito_data.get('ceiling_height', 'N/A'),
                        'Мебель': avito_data.get('furniture', 'N/A'),
                        'Способ продажи': avito_data.get('sale_type', 'N/A'),
                        'Просмотров сегодня': avito_data.get('today_views', 'N/A'),
                        'Адрес': avito_data.get('address', 'N/A'),
                        'Минут метро': avito_data.get('metro_time', 'N/A'),
                        'Метки': avito_data.get('tags', 'N/A'),
                        'Статус': 'Активно',
                        'Тип жилья': 'Квартира',
                    }
                    rows.append(excel_data)
                else:
                    print(f"❌ Не удалось спарсить объявление Avito: {url}")
            else:
                print(f"🏠 Парсим объявление Cian: {url}")
                # Для Cian используем существующую логику
                cian_data = parse_listing(url, sess)
                
                # Преобразуем данные Cian в формат для Excel
                if cian_data.get('total_floors') and cian_data['total_floors'] != 'N/A':
                    cian_data['Этаж'] = f"{cian_data.get('Этаж', 'N/A')}/{cian_data['total_floors']}"
                else:
                    cian_data['Этаж'] = cian_data.get('Этаж', 'N/A')
                
                rows.append(cian_data)
        except Exception as e:
            print(f"Ошибка при парсинге {url}: {e}")

    # Сохраняем и получаем request_id
    request_id = await save_listings(rows, user_id)

    # Формируем DataFrame
    df = pd.DataFrame(rows)

    # Обработка цен
    if 'Цена_raw' in df.columns:
        df['Цена'] = df['Цена_raw']
        df = df.sort_values('Цена_raw')
        df.drop('Цена_raw', axis=1, inplace=True)

    # Порядок колонок
    ordered = [
        'Комнат', 'Цена', 'Общая площадь', 'Жилая площадь',
        'Площадь кухни', 'Санузел', 'Балкон/лоджия', 'Вид из окон',
        'Ремонт', 'Этаж', 'Год постройки', 'Строительная серия',
        'Тип дома', 'Тип перекрытий', 'Пассажирских лифтов', 'Грузовых лифтов',
        'Парковка', 'Газоснабжение', 'Высота потолков', 'Мебель',
        'Способ продажи', 'Просмотров сегодня',
        'Адрес', 'Минут метро', 'Метки', 'Статус', 'Тип жилья', 'URL'
    ]
    df = df[[c for c in ordered if c in df.columns]]

    # Запись в BytesIO
    bio = BytesIO()
    df.to_excel(bio, index=False)
    bio.seek(0)

    # Если указан путь, сохраняем файл и форматируем столбец 'Цена'
    if output_path:
        with open(output_path, 'wb') as f:
            f.write(bio.getbuffer())

        wb = load_workbook(output_path)
        ws = wb.active
        # Выделяем заголовки
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Определяем колонку 'Цена' и задаем формат тысяч
        if 'Цена' in df.columns:
            price_idx = df.columns.get_loc('Цена') + 1
            price_col = get_column_letter(price_idx)
            custom_format = '#,##0'
            for row in range(2, ws.max_row + 1):
                cell = ws[f"{price_col}{row}"]
                if isinstance(cell.value, (int, float)):
                    cell.number_format = custom_format

        wb.save(output_path)

    return bio, request_id

# Полный парсинг страницы объявления
def parse_listing(url: str, session: requests.Session) -> dict:
    resp = session.get(url, headers=HEADERS)
    resp.encoding = 'utf-8'
    soup = BeautifulSoup(resp.text, 'html.parser')

    data = {'URL': url}
    # Определяем возможную блокировку (капча/антибот)
    page_text = soup.get_text(" ", strip=True).lower()
    is_blocked = bool(re.search(r"подтвердите, что запросы.*не робот|похожи на автоматические", page_text))
    if is_blocked:
        data['Статус'] = None
    elif soup.find(string=re.compile(r"Объявление снято", re.IGNORECASE)):
        data['Статус'] = 'Снято'
    labels = [span.get_text(strip=True) for span in soup.select('div[data-name="LabelsLayoutNew"] > span span:last-of-type')]
    data['Метки'] = '; '.join(labels) if labels else None
    h1 = soup.find('h1')
    if h1:
        m = re.search(r"(\d+)[^\d]*[-–]?комн", h1.get_text())
        if m:
            data['Комнат'] = extract_number(m.group(1))
    price_el = (
        soup.select_one('[data-name="NewbuildingPriceInfo"] [data-testid="price-amount"] span')
        or soup.select_one('[data-name="AsideGroup"] [data-testid="price-amount"] span')
    )
    if price_el:
        data['Цена_raw'] = extract_number(price_el.get_text())
        # Если статус ещё не определён и удалось найти цену — считаем объявление активным
        if 'Статус' not in data or data['Статус'] is None:
            data['Статус'] = 'Активно'

    summary = soup.select_one('[data-name="OfferSummaryInfoLayout"]')
    if summary:
        for item in summary.select('[data-name="OfferSummaryInfoItem"]'):
            ps = item.find_all('p')
            if len(ps) < 2:
                continue
            key = ps[0].get_text(strip=True)
            val = ps[1].get_text(strip=True)
            kl = key.lower().strip()
            if key == 'Строительная серия':
                data[key] = val
                continue
            if kl == 'этаж': 
                # Парсим этаж: "5 из 5" -> floor=5, total_floors=5
                floor_match = re.search(r'(\d+)\s*из\s*(\d+)', val)
                if floor_match:
                    data['Этаж'] = int(floor_match.group(1))
                    data['total_floors'] = int(floor_match.group(2))
                else:
                    data['Этаж'] = val
                    data['total_floors'] = None
                continue
            if kl in ['санузел', 'балкон/лоджия', 'количество лифтов']:
                data[key] = val; continue
            data[key] = extract_number(val) if re.search(r"\d", val) else val

    cont = soup.find('div', {'data-name': 'ObjectFactoids'})
    if cont:
        lines = cont.get_text(separator='\n', strip=True).split('\n')
        for i in range(0, len(lines)-1, 2):
            key, val = lines[i].strip(), lines[i+1].strip()
            kl = key.lower().strip()
            if key == 'Строительная серия': data[key] = val; continue
            if kl == 'этаж' and 'Этаж' not in data: 
                # Парсим этаж: "5 из 5" -> floor=5, total_floors=5
                floor_match = re.search(r'(\d+)\s*из\s*(\d+)', val)
                if floor_match:
                    data['Этаж'] = int(floor_match.group(1))
                    data['total_floors'] = int(floor_match.group(2))
                else:
                    data['Этаж'] = val
                    data['total_floors'] = None
            elif kl in ['санузел', 'балкон/лоджия', 'количество лифтов']: data[key] = val
            else: data[key] = extract_number(val) if re.search(r"\d", val) else val

    stats_re = re.compile(r"([\d\s]+)\sпросмотр\S*,\s*(\d+)\sза сегодня,\s*(\d+)\sуникаль", re.IGNORECASE)
    st = soup.find(string=stats_re)
    if st:
        m = stats_re.search(st)
        data['Всего просмотров'], data['Просмотров сегодня'], data['Уникальных просмотров'] = (
            extract_number(m.group(1)), extract_number(m.group(2)), extract_number(m.group(3))
        )

    geo = soup.select_one('div[data-name="Geo"]')
    if geo:
        span = geo.find('span', itemprop='name')
        addr = span['content'] if span and span.get('content') else ', '.join(
            a.get_text(strip=True) for a in geo.select('a[data-name="AddressItem"]')
        )
        parts = [s.strip() for s in addr.split(',') if s.strip()]
        data['Адрес'] = ', '.join(parts[-2:]) if len(parts) > 1 else addr
        stations = []
        for li in geo.select('ul[data-name="UndergroundList"] li[data-name="UndergroundItem"]'):
            st_el = li.find('a', href=True)
            tm_el = li.find('span', class_=re.compile(r".*underground_time.*"))
            if st_el and tm_el:
                name = st_el.get_text(strip=True)
                m = re.search(r"(\d+)", tm_el.get_text(strip=True))
                stations.append((name, int(m.group(1)) if m else None))
        if stations:
            station, time_to = min(stations, key=lambda x: x[1] or float('inf'))
            data['Минут метро'] = f"{time_to} {station}"

    return data

def extract_urls(raw_input: str) -> tuple[list[str], int]:
    """Извлекает URL из текста и определяет их источник"""
    urls = re.findall(r'https?://[^\s,;]+', raw_input)
    
    # Создаем экземпляр процессора для определения источника
    processor = ListingsProcessor()
    
    # Анализируем источники
    avito_count = 0
    cian_count = 0
    unknown_count = 0
    
    for url in urls:
        if processor.is_avito_url(url):
            avito_count += 1
        elif processor.is_cian_url(url):
            cian_count += 1
        else:
            unknown_count += 1
    
    # Выводим статистику по источникам
    if avito_count > 0 or cian_count > 0:
        print(f"🔍 Анализ ссылок:")
        if avito_count > 0:
            print(f"   🏠 Avito: {avito_count}")
        if cian_count > 0:
            print(f"   🏠 Cian: {cian_count}")
        if unknown_count > 0:
            print(f"   ⚠️ Неизвестные: {unknown_count}")
    
    return urls, len(urls)

# Создаем экземпляр класса для использования в других модулях
listings_processor = ListingsProcessor()

if __name__ == '__main__':
    user_id = 12345
    urls = ["https://www.cian.ru/sale/flat/318826533/"]
    excel = asyncio.run(export_listings_to_excel(urls, user_id, output_path="listings_numeric.xlsx"))
    print("listings_numeric.xlsx создан.")
