# text_handlers.py
from aiogram.types import Message, BufferedInputFile
from listings_processor import listings_processor, export_listings_to_excel, extract_urls
from db_handler import find_similar_ads_grouped
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

# Функции для форматирования данных
def format_date(date_value):
    """Форматирует дату из ISO формата в DD.MM.YYYY"""
    if not date_value:
        return ""
    
    try:
        # Если это строка в ISO формате
        if isinstance(date_value, str):
            from datetime import datetime
            # Парсим ISO формат
            if 'T' in date_value:
                dt = datetime.fromisoformat(date_value.replace('Z', '+00:00'))
            else:
                dt = datetime.fromisoformat(date_value)
        else:
            # Если это уже datetime объект
            dt = date_value
        
        # Форматируем в DD.MM.YYYY
        return dt.strftime("%d.%m.%Y")
    except Exception as e:
        print(f"Ошибка форматирования даты {date_value}: {e}")
        return str(date_value)

def format_boolean(bool_value):
    """Форматирует логическое значение в 'да' или пусто"""
    if bool_value is None:
        return ""
    
    # Преобразуем в строку и проверяем
    str_value = str(bool_value).lower().strip()
    
    # Проверяем различные варианты True
    if str_value in ['true', '1', 'yes', 'да', 'активно', 'active']:
        return "да"
    
    # Все остальное - пусто
    return ""

def extract_listing_comments(text: str, urls: list[str]) -> list[str]:
    """
    Извлекает комментарии между ссылками для каждого объявления
    Логика:
    - Текст до первой ссылки - это subtitle (общий комментарий)
    - Текст от первой до второй ссылки - комментарий к первому объявлению
    - Текст от второй до третьей ссылки - комментарий ко второму объявлению
    - И так далее
    """
    import re
    comments = []
    
    if not urls:
        return comments
    
    # Находим позиции всех ссылок в тексте
    url_positions = []
    for url in urls:
        pos = text.find(url)
        if pos != -1:
            url_positions.append((pos, url))
    
    # Сортируем по позиции
    url_positions.sort(key=lambda x: x[0])
    
    # Извлекаем комментарии для каждого объявления
    for i in range(len(url_positions)):
        current_pos, current_url = url_positions[i]
        
        if i == 0:
            # Для первого объявления - комментарий после первой ссылки
            # Ищем следующую ссылку или конец текста
            if i + 1 < len(url_positions):
                next_pos = url_positions[i + 1][0]
                comment = text[current_pos + len(current_url):next_pos].strip()
            else:
                # Если это единственная ссылка, берем текст после неё
                comment = text[current_pos + len(current_url):].strip()
        else:
            # Для остальных объявлений - текст между текущей и следующей ссылкой
            if i + 1 < len(url_positions):
                next_pos = url_positions[i + 1][0]
                comment = text[current_pos + len(current_url):next_pos].strip()
            else:
                # Для последнего объявления - текст после последней ссылки
                comment = text[current_pos + len(current_url):].strip()
        
        # Очищаем комментарий от лишних символов
        comment = comment.replace("подбор", "").replace("подбор-", "").strip()
        # Убираем цифры в начале комментария (ограничения фото)
        comment = re.sub(r'^\d+\s*', '', comment).strip()
        
        # Добавляем комментарий
        comments.append(comment)
    
    # Дополняем пустыми комментариями до нужного количества
    while len(comments) < len(urls):
        comments.append("")
    
    return comments

async def handle_text_message(message: Message):
    text = message.text.strip()

    # Проверка на запрос кабинета
    if "кабинет" in text.lower():
        user_id = message.from_user.id
        cabinet_url = f"https://mrealty.netlify.app/link?i={user_id}"
        await message.answer(f"🏢 Ваш личный кабинет: {cabinet_url}")
        return

    is_selection_request = "подбор" in text.lower()
    use_embedded = "подбор" in text.lower() and "подбор-" not in text.lower()  # "подбор" = встроенные, "подбор-" = обычные
    urls, url_count = extract_urls(text)
    
    # Сразу отправляем подтверждение получения ссылок
    if url_count == 0:
        await message.answer("📋 Отправьте ссылки на объявления CIAN для анализа.\n\n"
                             "💡 Для получения Excel-отчета просто отправьте ссылки.\n"
                             "🖼️ Для просмотра фотографий напишите 'подбор' + ссылки.\n"
                             "🔗 Для обычных ссылок: 'подбор-' + ссылки.")
        return
    
    await message.answer(f"✅ Принято ссылок: {url_count}")
    
    # Извлекаем метро из первого URL для названий файлов
    metro_info = "метро"
    # Убираем избыточный вызов extract_listing_info - он будет выполнен в generate_html_gallery_embedded
    if urls:
        try:
            # Простое извлечение метро из URL без полного парсинга
            if 'tekstilshchiki' in urls[0].lower() or 'текстильщики' in urls[0].lower():
                metro_info = "Текстильщики"
            elif 'begovaya' in urls[0].lower() or 'беговая' in urls[0].lower():
                metro_info = "Беговая"
            # Добавьте другие популярные станции метро при необходимости
        except Exception as e:
            print(f"Ошибка при извлечении метро: {e}")
            metro_info = "метро"

    try:
        if is_selection_request:
            subtitle = None
            max_photos_per_listing = None  # Максимальное количество фото на объявление
            listing_comments = []  # Комментарии к объявлениям
            
            if "подбор" in text.lower():
                podbor_pos = text.lower().find("подбор")
                
                # Ищем цифру перед словом "подбор"
                text_before_podbor = text[:podbor_pos].strip()
                import re
                
                # Ищем цифру в начале текста перед "подбор" (например: "3 подбор")
                print(f"🔍 DEBUG: Ищем ограничение фото в тексте: '{text_before_podbor}'")
                photo_limit_match = re.search(r'^(\d+)\s*', text_before_podbor)
                if photo_limit_match:
                    max_photos_per_listing = int(photo_limit_match.group(1))
                    print(f"🔢 Ограничение фото: {max_photos_per_listing} на объявление (найдено в начале)")
                else:
                    # Если не нашли в начале, ищем любую цифру перед "подбор"
                    photo_limit_match = re.search(r'(\d+)\s*$', text_before_podbor)
                    if photo_limit_match:
                        max_photos_per_listing = int(photo_limit_match.group(1))
                        print(f"🔢 Ограничение фото: {max_photos_per_listing} на объявление (найдено в конце)")
                    else:
                        print(f"🔍 DEBUG: Ограничение фото не найдено, будут загружены все доступные фото")
                
                text_after_podbor = text[podbor_pos + 6:].strip()
                first_url_pos = -1
                for url in urls:
                    pos = text_after_podbor.find(url)
                    if pos != -1:
                        first_url_pos = pos
                        break
                if first_url_pos != -1:
                    subtitle = text_after_podbor[:first_url_pos].strip()
                else:
                    subtitle = text_after_podbor
                if subtitle:
                    subtitle = subtitle.replace("подбор", "").replace("подбор-", "").strip()
                
                # Извлекаем комментарии между ссылками
                listing_comments = extract_listing_comments(text, urls)
                print(f"📝 Найдено комментариев к объявлениям: {len(listing_comments)}")

            if use_embedded:
                html_content, photo_stats = await listings_processor.generate_html_gallery_embedded(urls, message.from_user.id, subtitle, remove_watermarks=True, max_photos_per_listing=max_photos_per_listing, listing_comments=listing_comments)
                filename = f"Подбор_{metro_info}.html"
                caption = f"🏠 Подбор недвижимости"
                
                # Показываем статус загрузки фото для каждого объявления
                if max_photos_per_listing:
                    await message.answer(f"🔢 Ограничение: максимум {max_photos_per_listing} фото на объявление")
                
                for stat in photo_stats:
                    if stat['photo_count'] > 0:
                        await message.answer(f"📸 {stat['photo_count']} фото загружено из объявления {stat['listing_number']}")
                    else:
                        if 'error' in stat:
                            await message.answer(f"❌ Ошибка при загрузке фото из объявления {stat['listing_number']}: {stat['error']}")
                        else:
                            await message.answer(f"⚠️ Фото не найдены в объявлении {stat['listing_number']}")
            else:
                html_content = await listings_processor.generate_html_gallery(urls, message.from_user.id, subtitle, listing_comments, max_photos_per_listing)
                filename = f"Подбор_{metro_info}.html"
                caption = f"🏠 Подбор недвижимости (обычные ссылки)"

            html_file = io.BytesIO(html_content.encode('utf-8'))
            html_file.name = filename
            input_file = BufferedInputFile(html_file.getvalue(), filename=filename)
            await message.answer_document(input_file, caption=caption)

        else:
            excel_file, request_id = await export_listings_to_excel(urls, message.from_user.id)
            input_file = BufferedInputFile(excel_file.getvalue(), filename=f"Сравнение_{metro_info}.xlsx")
            await message.answer_document(input_file, caption=f"📊 Сравнение недвижимости")
            
            # Создаем дополнительный Excel файл с похожими объявлениями
            try:
                similar_ads = await find_similar_ads_grouped(request_id)
                if similar_ads:
                    # Создаем Excel файл с похожими объявлениями
                    similar_excel = io.BytesIO()
                    
                    # Создаем DataFrame для похожих объявлений
                    import pandas as pd
                    import json
                    similar_data = []
                    for group in similar_ads:
                        address = group['address']
                        ads = group['ads']
                        
                        # Проверяем, что ads - это список, а не строка
                        if isinstance(ads, list):
                            ads_list = ads
                        elif isinstance(ads, str):
                            try:
                                # Пытаемся распарсить JSON строку
                                ads_list = json.loads(ads)
                                if not isinstance(ads_list, list):
                                    print(f"⚠️ ads после парсинга не является списком: {type(ads_list)}")
                                    continue
                            except json.JSONDecodeError as e:
                                print(f"⚠️ Ошибка парсинга JSON: {e}")
                                continue
                        else:
                            print(f"⚠️ ads не является списком или строкой: {type(ads)}")
                            continue
                        
                        # Обрабатываем список объявлений
                        for ad in ads_list:
                            if isinstance(ad, dict):
                                ad_copy = ad.copy()  # Создаем копию, чтобы не изменять оригинал
                                ad_copy['Адрес'] = address
                                
                                # Форматируем даты и логические значения
                                if 'created' in ad_copy:
                                    ad_copy['created'] = format_date(ad_copy['created'])
                                if 'updated' in ad_copy:
                                    ad_copy['updated'] = format_date(ad_copy['updated'])
                                if 'is_active' in ad_copy:
                                    ad_copy['is_active'] = format_boolean(ad_copy['is_active'])
                                
                                similar_data.append(ad_copy)
                    
                    if similar_data:
                        # Создаем Excel файл с группировкой по адресам
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Похожие объявления"
                        
                        # Заголовки для объявлений
                        headers = ['URL', 'Цена', 'Комнат', 'Создано', 'Обновлено', 'Активно', 'Тип продавца']
                        header_row = 1
                        
                        # Группируем данные по адресам
                        current_row = 1
                        for group in similar_ads:
                            address = group['address']
                            ads = group['ads']
                            
                            # Парсим ads если это строка
                            if isinstance(ads, str):
                                try:
                                    ads_list = json.loads(ads)
                                except json.JSONDecodeError:
                                    continue
                            else:
                                ads_list = ads
                            
                            if not isinstance(ads_list, list):
                                continue
                            
                            # Добавляем адрес как заголовок (жирным шрифтом)
                            ws.cell(row=current_row, column=1, value=address)
                            ws.cell(row=current_row, column=1).font = Font(bold=True)
                            current_row += 1
                            
                            # Добавляем заголовки колонок
                            for col, header in enumerate(headers, 1):
                                ws.cell(row=current_row, column=col, value=header)
                                ws.cell(row=current_row, column=col).font = Font(bold=True)
                            current_row += 1
                            
                            # Добавляем объявления
                            for ad in ads_list:
                                if isinstance(ad, dict):
                                    ws.cell(row=current_row, column=1, value=ad.get('url', ''))
                                    ws.cell(row=current_row, column=2, value=ad.get('price', ''))
                                    ws.cell(row=current_row, column=3, value=ad.get('rooms', ''))
                                    # Форматируем даты
                                    ws.cell(row=current_row, column=4, value=format_date(ad.get('created', '')))
                                    ws.cell(row=current_row, column=5, value=format_date(ad.get('updated', '')))
                                    # Форматируем логическое значение
                                    ws.cell(row=current_row, column=6, value=format_boolean(ad.get('is_active', '')))
                                    ws.cell(row=current_row, column=7, value=ad.get('person_type', ''))
                                    current_row += 1
                            
                            # Добавляем пустую строку между группами
                            current_row += 1
                        
                        # Сохраняем в BytesIO
                        similar_excel = io.BytesIO()
                        wb.save(similar_excel)
                        similar_excel.seek(0)
                        
                        similar_filename = f"похожие_объявления_{metro_info}.xlsx"
                        similar_input_file = BufferedInputFile(similar_excel.getvalue(), filename=similar_filename)
                        await message.answer_document(similar_input_file, caption=f"🔍 Похожие объявления")
                        print(f"✅ Создан Excel с {len(similar_data)} похожими объявлениями")
                    else:
                        print("⚠️ Нет данных для создания Excel с похожими объявлениями")
            except Exception as e:
                print(f"Ошибка при создании файла с похожими объявлениями: {e}")
                # Не отправляем сообщение об ошибке пользователю, чтобы не засорять чат

    except Exception as e:
        await message.answer(f"❌ Ошибка при обработке: {str(e)}\n\nПопробуйте еще раз или обратитесь к администратору.")
        print(f"Ошибка в handle_text_message: {e}")
